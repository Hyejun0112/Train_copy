import os
import sys
import subprocess
import csv
import math
import re
import shutil
import tempfile
import traceback
import time
import threading
import itertools
import multiprocessing as mp
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pyautogui
import pygetwindow as gw

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

# ── 전역 설정 ──────────────────────────────────────────────
WAIT_OPEN  = 5
WAIT_SHORT = 0.8
WAIT_PASTE = 2.5

_APP_TITLE = "Bluebeam Markup Auto-Copy"  # 우리 GUI 창 제목(close_bluebeam_app에서 자기 자신 제외용)

# ── 상태 변수 ──────────────────────────────────────────────
src_folder    = ""
dst_folder    = ""
output_folder = ""
src_files     = []
dst_files     = []
mapping       = []
stop_flag     = False

# 마크업 필터 설정
filter_enabled   = False
filter_color     = ""   # RGB hex, 예: "0000FF"
filter_date_limit = ""  # "YYYY-MM-DD" (이 날짜까지 작성된 마크업만 포함)
filter_author    = ""   # 작성자(사번)

# 위치 보정 설정 (도면 레이아웃이 다른 경우) — Train 번호만 다른 동일 Tag 자동 매칭으로 기준점 계산
pos_correction_enabled = False


def _list_pdfs(path: str) -> list:
    """폴더 내 PDF 목록 (필터 작업용 임시 파일 '_filtered_*'는 제외)"""
    return [
        f for f in os.listdir(path)
        if f.lower().endswith(".pdf") and not f.startswith("_filtered_")
    ]



# ══════════════════════════════════════════════════════════
#  Bluebeam 제어 유틸
# ══════════════════════════════════════════════════════════

def focus_bluebeam(timeout: int = 10) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        wins = gw.getWindowsWithTitle("Bluebeam")
        if wins:
            try:
                wins[0].activate()
                time.sleep(0.8)
                return True
            except Exception:
                pass
        time.sleep(0.5)
    return False


def fit_page():
    pyautogui.hotkey('ctrl', 'shift', 'h')
    time.sleep(WAIT_SHORT)


def open_pdf(path: str):
    os.startfile(path)
    time.sleep(WAIT_OPEN)
    focus_bluebeam()
    fit_page()


def open_folder(path: str) -> bool:
    """탐색기/파인더로 폴더를 연다. 성공 여부를 반환(실패해도 예외는 안 던짐)."""
    if not path or not os.path.isdir(path):
        return False
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
        return True
    except Exception:
        return False


def close_pdf_discard():
    """저장 없이 닫기 — 저장 여부 팝업이 뜬 경우에만 N으로 닫음
    (팝업이 없을 때 무조건 'n'을 누르면 Bluebeam의 Note 도구 단축키(N)가
    실행되어 다음 작업 중 의도치 않은 Note 마크업이 생성되는 문제가 있었음)"""
    pyautogui.hotkey('ctrl', 'w')
    time.sleep(WAIT_SHORT)
    try:
        active = gw.getActiveWindow()
        title = (active.title if active else "") or ""
    except Exception:
        title = ""
    if title and "Bluebeam" not in title:
        pyautogui.press('n')
        time.sleep(WAIT_SHORT)


def close_bluebeam_app():
    """현재 Bluebeam 창을 완전히 종료한다.
    - 전체화면 상태로 남아 응답 없음(렉)이 발생하는 것을 방지하기 위해 먼저 Escape로 빠져나옴
    - 작업이 끝난 PDF 창이 계속 쌓이지 않도록 Alt+F4로 앱 전체를 닫음
    - 완전히 닫힐 때까지 최대 15초 대기
    주의: 우리 GUI 창 제목이 "Bluebeam Markup Auto-Copy"라서 단순히 제목에
    "Bluebeam"이 들어간 창을 찾으면 우리 자신의 창도 걸려서 Alt+F4로 우리
    프로그램이 종료되어 버린다(위치 보정 모드는 Bluebeam을 전혀 안 쓰는데도
    이 부작용으로 작업 완료 후 프로그램이 꺼지는 문제가 있었음). 우리 창은
    제외하고 진짜 Bluebeam 창만 대상으로 한다."""
    pyautogui.press('escape')
    time.sleep(WAIT_SHORT)

    wins = [w for w in gw.getWindowsWithTitle("Bluebeam")
            if w.title != _APP_TITLE]
    if not wins:
        return
    try:
        wins[0].activate()
        time.sleep(0.3)
    except Exception:
        pass

    pyautogui.hotkey('alt', 'f4')
    time.sleep(WAIT_SHORT)

    # 저장 확인 팝업 처리 (변경사항 있을 경우)
    try:
        active = gw.getActiveWindow()
        title = (active.title if active else "") or ""
    except Exception:
        title = ""
    if title and "Bluebeam" not in title:
        pyautogui.press('n')
        time.sleep(WAIT_SHORT)

    # Bluebeam이 완전히 닫힐 때까지 최대 15초 대기
    for _ in range(30):
        time.sleep(0.5)
        try:
            remaining = gw.getWindowsWithTitle("Bluebeam")
        except Exception:
            remaining = []
        if not remaining:
            break


# ══════════════════════════════════════════════════════════
#  핵심 작업 로직
# ══════════════════════════════════════════════════════════

class StoppedError(Exception):
    pass


def _parse_pdf_date(date_str):
    """PDF 날짜 형식 'D:YYYYMMDDHHMMSS...' → 'YYYY-MM-DD' (실패 시 None)"""
    if not date_str:
        return None
    s = date_str.lstrip("D:")
    if len(s) < 8:
        return None
    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"


def _normalize_date_input(date_str):
    """사용자 입력 날짜를 'YYYY-MM-DD'로 정규화 (2026.05.08, 2026/5/8 등 허용, 실패 시 "")"""
    s = (date_str or "").strip()
    if not s:
        return ""
    parts = s.replace(".", "-").replace("/", "-").split("-")
    if len(parts) != 3:
        return ""
    try:
        y, m, d = (int(p) for p in parts)
        return f"{y:04d}-{m:02d}-{d:02d}"
    except ValueError:
        return ""


def _get_font_color(annot):
    """FreeText(텍스트) 마크업의 글꼴 색상을 /DA 문자열에서 추출 (RGB 0~1 튜플, 실패 시 None)"""
    try:
        da = annot.parent.xref_get_key(annot.xref, "DA")
    except Exception:
        return None
    if not da or da[0] != "string":
        return None
    parts = da[1].split()
    for i, tok in enumerate(parts):
        if tok == "rg" and i >= 3:
            try:
                return tuple(float(parts[i - 3 + j]) for j in range(3))
            except (ValueError, IndexError):
                return None
        if tok == "g" and i >= 1:
            try:
                v = float(parts[i - 1])
                return (v, v, v)
            except (ValueError, IndexError):
                return None
    return None


def _annot_matches_filter(annot, color_hex, date_limit, author):
    """주어진 조건을 모두 만족해야 True (조건이 비어있으면 해당 항목은 통과)"""
    if color_hex:
        colors = annot.colors or {}
        target = tuple(int(color_hex[i:i + 2], 16) / 255 for i in (0, 2, 4))

        def _close(c):
            return bool(c) and len(c) == 3 and all(
                abs(a - b) <= 0.15 for a, b in zip(c, target)
            )

        # FreeText(텍스트) 마크업: "글꼴 색상" 기준
        # 그 외 마크업: "색" (테두리/기본색, stroke) 기준
        if annot.type[1] == "FreeText":
            font_color = _get_font_color(annot)
            # 콜아웃형 FreeText는 /DA의 글꼴색이 실제 표시 색(테두리/채우기)과
            # 다를 수 있으므로, 글꼴 색·테두리색·채우기색 중 하나라도 맞으면 통과
            if not (
                _close(font_color) or
                _close(colors.get("stroke")) or
                _close(colors.get("fill"))
            ):
                return False
        else:
            if not (_close(colors.get("stroke")) or _close(colors.get("fill"))):
                return False

    if author:
        annot_author = (annot.info.get("title") or "").strip()
        if not annot_author:
            # 일부 마크업 유형은 annot.info에 title이 채워지지 않으므로
            # PDF의 /T 키를 직접 읽어 대체 확인
            try:
                kind, val = annot.parent.xref_get_key(annot.xref, "T")
                if kind == "string":
                    annot_author = val.strip()
            except Exception:
                pass
        if annot_author.lower() != author.strip().lower():
            return False

    if date_limit:
        # 마크업이 "작성된" 날짜(creationDate) 기준, 없으면 수정일(modDate)로 대체
        d = _parse_pdf_date(annot.info.get("creationDate") or annot.info.get("modDate"))
        if d is None:
            return False
        if d > date_limit:
            return False

    return True


def build_filtered_copy(src_path, color_hex, date_limit, author, log_fn=None):
    """필터 조건에 맞지 않는 마크업(annotation)을 제거한 임시 PDF를 만들어
    그 경로를 반환한다. fitz(PyMuPDF)가 없으면 RuntimeError."""
    if fitz is None:
        raise RuntimeError("PyMuPDF(fitz)가 설치되어 있지 않습니다. 'pip install PyMuPDF' 필요")

    doc = fitz.open(src_path)
    kept, removed = 0, 0
    authors_seen = set()
    for page in doc:
        for annot in page.annots() or []:
            a = (annot.info.get("title") or "").strip()
            if not a:
                try:
                    kind, val = annot.parent.xref_get_key(annot.xref, "T")
                    if kind == "string":
                        a = val.strip()
                except Exception:
                    pass
            if a:
                authors_seen.add(a)

        # 삭제할 annot의 xref를 하나씩 찾아 삭제 -> 매번 처음부터 다시 스캔
        # (delete_annot 호출 후에는 기존에 모아둔 annot/xref가 무효화될 수 있음)
        while True:
            target_xref = None
            for annot in page.annots() or []:
                if _annot_matches_filter(annot, color_hex, date_limit, author):
                    continue
                target_xref = annot.xref
                break

            if target_xref is None:
                break

            annot = page.load_annot(target_xref)
            page.delete_annot(annot)
            removed += 1

        for annot in page.annots() or []:
            kept += 1

    if log_fn:
        log_fn(f"  [필터] 유지 {kept}개 / 제외 {removed}개\n")
        if author:
            log_fn(f"  [필터] 발견된 작성자 목록: {sorted(authors_seen)}\n")

    fd, tmp_path = tempfile.mkstemp(
        suffix=".pdf", prefix="_filtered_", dir=tempfile.gettempdir()
    )
    os.close(fd)
    doc.save(tmp_path)
    doc.close()
    return tmp_path, kept, removed


# ══════════════════════════════════════════════════════════
#  위치 보정 (도면 레이아웃이 다른 경우) — PyMuPDF 직접 복사
# ══════════════════════════════════════════════════════════

# Train 번호(앞쪽 2~4자리 숫자)만 다르고 나머지는 동일한 Tag 패턴
# 예: "504-CWS-0100-400-ACB3B02SE51-NN" / "604-CWS-0100-400-ACB3B02SE51-NN"
_TAG_RE = re.compile(r'^(\d{2,4})-([A-Za-z0-9][A-Za-z0-9-]{4,})$')


def _extract_tag_suffixes(page):
    """페이지에서 'NNN-나머지' 형태의 Tag 텍스트를 추출해 suffix(나머지 부분)별
    위치(중심점)를 반환. 같은 suffix가 페이지에 여러 번 나오면 어느 게 맞는지
    모호하므로 매칭에서 제외한다."""
    suffix_map = {}
    try:
        words = page.get_text("words")
    except Exception:
        return {}
    for w in words:
        x0, y0, x1, y1, text = w[0], w[1], w[2], w[3], w[4]
        m = _TAG_RE.match(text)
        if not m:
            continue
        suffix = m.group(2)
        center = ((x0 + x1) / 2, (y0 + y1) / 2)
        if suffix in suffix_map:
            suffix_map[suffix] = None  # 중복 발견 → 모호함 표시
        else:
            suffix_map[suffix] = center
    return {k: v for k, v in suffix_map.items() if v is not None}


# 계측기/제어밸브/On-off 밸브/Logic/OPC 등 일반 Tag (예: "PI-0101", "CV-0022",
# "XV-0011", "LSH-101", "B-101") — Train 번호와 무관하게 Source/Target에서
# 보통 동일한 문자가 그대로 유지된다. 이런 Tag는 Train 번호 Tag(배관선번호)보다
# 도면 전체에 훨씬 빽빽하게 분포하므로, 마크업 바로 옆의 동일 Tag를 기준점으로
# 쓰면 국소 보정 정확도가 올라간다.
# 단, "수동(Manual) 밸브" Tag(GV, BFV, BV, GLV, PLV, NRV, CKV, DV 등)는
# Train Copy마다 번호가 별도로 매겨지므로 같은 번호라도 다른 설비를 가리킬 수
# 있어 기준점에서 제외해야 한다. 반면 Control Valve(CV)나 On/off Valve(XV, SOV,
# AOV, MOV 등)는 Loop/Logic 번호에 종속돼 있어 그대로 유지되므로 사용 가능하다.
_GENERIC_TAG_RE = re.compile(r'^[A-Za-z]{1,6}-\d{2,6}[A-Za-z0-9]*$')
_MANUAL_VALVE_PREFIX_RE = re.compile(
    r'^(?:GLV|PLV|NRV|CKV|BFV|GV|BV|DV)-', re.IGNORECASE
)


def _extract_generic_tags(page):
    """페이지에서 일반 계측/제어밸브/Logic/OPC Tag 텍스트를 추출해 텍스트별
    위치(중심점)를 반환. 수동 밸브 Tag는 Train Copy마다 번호가 달라질 수 있어
    제외하고, 같은 텍스트가 페이지에 여러 번 나오면 모호하므로 매칭에서 제외한다."""
    text_map = {}
    try:
        words = page.get_text("words")
    except Exception:
        return {}
    for w in words:
        x0, y0, x1, y1, text = w[0], w[1], w[2], w[3], w[4]
        if not _GENERIC_TAG_RE.match(text):
            continue
        if _MANUAL_VALVE_PREFIX_RE.match(text):
            continue
        center = ((x0 + x1) / 2, (y0 + y1) / 2)
        if text in text_map:
            text_map[text] = None  # 중복 발견 → 모호함 표시
        else:
            text_map[text] = center
    return {k: v for k, v in text_map.items() if v is not None}


# Symbol(벡터 도형) 기준점: 작은 벡터 도형 묶음(밸브/계측기 심볼 등)을 묶어
# (도형 path 개수, bbox 가로/세로 크기)로 모양 시그니처를 만들고, 그 시그니처가
# 페이지에 단 한 번만 나오는 경우만 기준점으로 사용한다(Tag 텍스트가 없는
# 심볼도 위치 보정에 활용 가능).
_SYMBOL_MAX_SIZE = 60  # pt — 이보다 크면 개별 심볼이 아닌 배경/라인 묶음으로 간주해 제외


def _extract_symbol_signatures(page):
    try:
        drawings = page.get_drawings()
    except Exception:
        return {}
    sig_map = {}
    for d in drawings:
        rect = d.get("rect")
        if rect is None or rect.width <= 0 or rect.height <= 0:
            continue
        if rect.width > _SYMBOL_MAX_SIZE or rect.height > _SYMBOL_MAX_SIZE:
            continue
        items = d.get("items", [])
        sig = (len(items), round(rect.width, 1), round(rect.height, 1))
        center = ((rect.x0 + rect.x1) / 2, (rect.y0 + rect.y1) / 2)
        if sig in sig_map:
            sig_map[sig] = None
        else:
            sig_map[sig] = center
    return {k: v for k, v in sig_map.items() if v is not None}


def _border_corner_anchors(src_page, dst_page):
    """Source/Target 페이지 크기를 기준점(네 모서리)으로 추가한다. 도면 안쪽에
    매칭된 Tag/Symbol이 없는 외곽 영역의 마크업에 대한 안전한 보조 기준점
    역할을 한다(Source/Target 페이지 크기가 같으면 이동량 0으로 반영됨)."""
    sr, dr = src_page.rect, dst_page.rect
    s_corners = [(sr.x0, sr.y0), (sr.x1, sr.y0), (sr.x0, sr.y1), (sr.x1, sr.y1)]
    d_corners = [(dr.x0, dr.y0), (dr.x1, dr.y0), (dr.x0, dr.y1), (dr.x1, dr.y1)]
    return list(zip(s_corners, d_corners))


# 사용자가 Bluebeam에서 직접 찍어두는 "수동 기준점" 마크업 색상. 이 색(마젠타)으로
# 그려진 마크업을 Source/Target에서 각각 찾아, 같은 순서로 1:1 매칭해 가장 신뢰도
# 높은 기준점으로 사용한다. 자동 Tag/Symbol 매칭이 오매칭될 위험이 있는 도면에서,
# 사람이 눈으로 확인한 동일 위치를 보장하기 위함. 이 마크업 자체는 복사 대상에서 제외.
_MANUAL_ANCHOR_COLOR = (1.0, 0.0, 1.0)
_MANUAL_ANCHOR_TOL = 0.08


def _is_anchor_color(annot) -> bool:
    colors = annot.colors or {}
    for c in (colors.get("stroke"), colors.get("fill")):
        if c and len(c) == 3 and all(
            abs(c[i] - _MANUAL_ANCHOR_COLOR[i]) < _MANUAL_ANCHOR_TOL for i in range(3)
        ):
            return True
    return False


def _extract_manual_anchor_points(page):
    """페이지에서 마젠타색 기준점 마크업을 찾아 (위치 순서대로 정렬된 중심점 목록,
    해당 마크업들의 xref 집합)을 반환한다. xref는 복사 시 제외하기 위함."""
    found = []
    for annot in page.annots() or []:
        if not _is_anchor_color(annot):
            continue
        r = annot.rect
        found.append((annot.xref, ((r.x0 + r.x1) / 2, (r.y0 + r.y1) / 2)))
    return found


def _find_tag_matches(src_doc, dst_doc, log_fn=None):
    """Source/Target 문서에서 기준점을 찾아 매칭한다. 우선순위:
    (0) 사용자가 마젠타색으로 직접 찍어둔 수동 기준점 마크업(있으면 이것만 사용 —
        가장 신뢰도 높음), 없으면 자동으로 (1) Train 번호만 다른 동일 Tag(배관선번호
        suffix), (2) 텍스트가 완전히 동일한 일반 계측/제어밸브/Logic/OPC Tag,
        (3) 모양이 동일한 Symbol(벡터 도형)을 찾아 매칭하고, (4) 페이지 모서리를
        보조 기준점으로 추가한다.
    반환: (src_page_idx, dst_page_idx, [(src_pt, dst_pt), ...], skip_src_xrefs,
          [(tag_name, src_pt, dst_pt), ...]) — 못 찾으면 None.
    마지막 named_pairs는 마크업이 '소속된 Tag'를 찾기 위한 용도로, 모서리
    보조점처럼 이름이 없는 anchor는 포함하지 않는다."""
    def log(msg):
        if log_fn:
            log_fn(msg)

    page_pairs = list(itertools.product(src_doc, dst_doc))

    for sp, dp in page_pairs:
        src_anchors = _extract_manual_anchor_points(sp)
        dst_anchors = _extract_manual_anchor_points(dp)
        if src_anchors and dst_anchors and len(src_anchors) == len(dst_anchors):
            src_sorted = sorted(src_anchors, key=lambda t: t[0])
            dst_sorted = sorted(dst_anchors, key=lambda t: t[0])
            pairs = [(s[1], d[1]) for s, d in zip(src_sorted, dst_sorted)]
            skip_xrefs = {xref for xref, _ in src_sorted}
            log(f"  [위치 보정] 수동 기준점(마젠타 마크업) {len(pairs)}개 발견 — "
                f"이 기준점만 사용 (Source p{sp.number + 1} → Target p{dp.number + 1})\n")
            return sp.number, dp.number, pairs, skip_xrefs, []

    for sp, dp in page_pairs:
        src_map = _extract_tag_suffixes(sp)
        src_generic = _extract_generic_tags(sp)
        src_symbols = _extract_symbol_signatures(sp)
        src_bubbles = _extract_instrument_bubbles(sp)
        if not src_map and not src_generic and not src_symbols and not src_bubbles:
            continue
        dst_map = _extract_tag_suffixes(dp)
        dst_generic = _extract_generic_tags(dp)
        dst_symbols = _extract_symbol_signatures(dp)
        dst_bubbles = _extract_instrument_bubbles(dp)
        common = sorted(set(src_map) & set(dst_map))
        common_generic = sorted(set(src_generic) & set(dst_generic))
        common_symbols = sorted(set(src_symbols) & set(dst_symbols))
        common_bubbles = sorted(set(src_bubbles) & set(dst_bubbles))
        total = len(common) + len(common_generic) + len(common_symbols) + len(common_bubbles)
        if total >= 2:
            # named_pairs: 마크업이 '소속된 Tag'를 찾을 수 있도록 Tag 텍스트와
            # 함께 보관한다(아래 _cluster_tag_anchor에서 사용). 모서리 보조점은
            # 이름이 없는 anchor라 named_pairs에서 제외한다.
            named_pairs = [(s, src_map[s], dst_map[s]) for s in common]
            named_pairs += [(s, src_generic[s], dst_generic[s]) for s in common_generic]
            named_pairs += [(s, src_symbols[s], dst_symbols[s]) for s in common_symbols]
            named_pairs += [(s, src_bubbles[s], dst_bubbles[s]) for s in common_bubbles]
            pairs = [(sp_, dp_) for _, sp_, dp_ in named_pairs]
            pairs += _border_corner_anchors(sp, dp)
            log(f"  [위치 보정] 자동 기준점 매칭 {total}개 발견 "
                f"(배관선번호 {len(common)}개 + 일반 Tag {len(common_generic)}개 + "
                f"Symbol {len(common_symbols)}개 + 계기 버블 {len(common_bubbles)}개, "
                f"모서리 4개 보조 추가, "
                f"Source p{sp.number + 1} → Target p{dp.number + 1})\n")
            for s in common[:10]:
                log(f"    - {s}\n")
            for s in common_generic[:10]:
                log(f"    - {s}\n")
            for s in common_bubbles[:10]:
                log(f"    - {s}\n")
            return sp.number, dp.number, pairs, set(), named_pairs
    return None


def _local_offset_matrix(pt, pairs, k=3):
    """pt(소스 좌표) 근처의 매칭된 Tag k개(거리 역제곱 가중)를 이용해
    그 지역의 회전/스케일/이동을 모두 반영하는 유사변환(similarity transform)
    행렬을 만든다. 기준점이 1개뿐이면 평행이동만 적용한다.
    도면이 전체적으로 하나의 회전/스케일로 맞지 않고 구역마다 CAD 배치가
    달라진 경우, 전역 변환보다 마크업 주변의 국소 변환을 쓰는 게 더 정확하다."""
    p = complex(*pt)
    dists = sorted(
        ((abs(complex(*sp) - p), sp, dp) for sp, dp in pairs),
        key=lambda t: t[0]
    )
    chosen = dists[:max(1, min(k, len(dists)))]
    weights = [1.0 / (d * d + 1.0) for d, _, _ in chosen]
    wsum = sum(weights)

    zs = [complex(*sp) for _, sp, _ in chosen]
    ws = [complex(*dp) for _, _, dp in chosen]
    zbar = sum(w * z for w, z in zip(weights, zs)) / wsum
    wbar = sum(w * v for w, v in zip(weights, ws)) / wsum

    den = sum(wt * abs(z - zbar) ** 2 for wt, z in zip(weights, zs))
    if den < 1e-9 or len(chosen) < 2:
        # 기준점이 1개거나 모두 같은 위치면 회전/스케일을 구할 수 없어 평행이동만 적용
        dx = wbar.real - zbar.real
        dy = wbar.imag - zbar.imag
        return fitz.Matrix(1, 0, 0, 1, dx, dy)

    num = sum(
        wt * (z - zbar).conjugate() * (v - wbar)
        for wt, z, v in zip(weights, zs, ws)
    )
    a = num / den  # 복소수: 회전+스케일을 동시에 표현
    scale = abs(a)

    # 가까운 기준점들이 한 줄로 거의 모여있으면(분모가 작음) 회전/스케일 추정이
    # 불안정해져 마크업이 비정상적으로 늘어나거나 뒤집힐 수 있다.
    # 스케일이 비정상적인 범위면 평행이동만 적용하는 안전한 방식으로 되돌린다.
    if not math.isfinite(scale) or scale < 0.5 or scale > 2.0:
        dx = wbar.real - zbar.real
        dy = wbar.imag - zbar.imag
        return fitz.Matrix(1, 0, 0, 1, dx, dy)

    b = wbar - a * zbar
    return fitz.Matrix(a.real, a.imag, -a.imag, a.real, b.real, b.imag)


def _global_similarity_matrix(pairs):
    """모든 기준점 쌍으로 도면 한 장 전체에 적용할 단일 유사변환(회전+스케일+
    이동)을 최소제곱 적합한다. 마크업마다 다른 국소 변환을 쓰면 클라우드+콜아웃
    +지시선처럼 여러 조각으로 이루어진 마크업이 서로 어긋나(따로 놀고) 보이므로,
    한 장 전체에 동일한 변환을 적용해 마크업 간 상대 위치를 그대로 보존한다.
    유사변환은 평행/직각/비율을 보존하므로 기울어짐(전단)도 생기지 않는다."""
    n = len(pairs)
    if n == 0:
        return fitz.Matrix(1, 0, 0, 1, 0, 0)
    zs = [complex(*sp) for sp, _ in pairs]
    ws = [complex(*dp) for _, dp in pairs]
    zbar = sum(zs) / n
    wbar = sum(ws) / n
    den = sum(abs(z - zbar) ** 2 for z in zs)
    if den < 1e-9 or n < 2:
        return fitz.Matrix(1, 0, 0, 1, wbar.real - zbar.real, wbar.imag - zbar.imag)
    num = sum((z - zbar).conjugate() * (v - wbar) for z, v in zip(zs, ws))
    a = num / den
    scale = abs(a)
    if not math.isfinite(scale) or scale < 0.5 or scale > 2.0:
        return fitz.Matrix(1, 0, 0, 1, wbar.real - zbar.real, wbar.imag - zbar.imag)
    b = wbar - a * zbar
    return fitz.Matrix(a.real, a.imag, -a.imag, a.real, b.real, b.imag)


def _fit_scale_translate(pairs):
    """주어진 기준점 쌍으로 '회전 없는' 균일 스케일+평행이동을 최소제곱 적합해
    (s, sx, sy, dx, dy)를 돌려준다. den이 0이면 s=None."""
    n = len(pairs)
    sx = sum(sp[0] for sp, _ in pairs) / n
    sy = sum(sp[1] for sp, _ in pairs) / n
    dx = sum(dp[0] for _, dp in pairs) / n
    dy = sum(dp[1] for _, dp in pairs) / n
    num = sum((sp[0] - sx) * (dp[0] - dx) + (sp[1] - sy) * (dp[1] - dy)
              for sp, dp in pairs)
    den = sum((sp[0] - sx) ** 2 + (sp[1] - sy) ** 2 for sp, _ in pairs)
    if den < 1e-9 or n < 2:
        return None, sx, sy, dx, dy
    return num / den, sx, sy, dx, dy


def _global_scale_translate_matrix(pairs):
    """모든 기준점 쌍으로 '회전 없는' 단일 변환(균일 스케일 + 평행이동, 자유도 3)을
    적합한다. 단순 최소제곱은 오매칭된 기준점 하나가 배율을 통째로 끌어당겨(예:
    같은 도면 템플릿인데도 배율 0.90), 그 잘못된 축소가 모든 마크업에 균일하게
    적용돼 이미 잘 맞던 영역까지 틀어뜨린다. → '이상치에 강인한' 절사
    최소제곱으로 바꾼다: (1) 강인한 중앙값 평행이동을 씨앗으로 잡고 (2) 그
    평행이동 기준 잔차가 큰 상위 30%(오매칭·국소 이동 설비)를 버린 뒤 (3) 남은
    내부값(inlier)으로만 배율을 다시 적합한다. 같은 시트 템플릿이라 정상 배율은
    1.0 근처여야 하므로, 적합 배율이 ±5%를 벗어나면 오매칭 잔재로 보고 순수
    평행이동(배율 1.0)으로 되돌린다. 남는 국소 오차는 _idw_offset이 따로 잡는다."""
    n = len(pairs)
    if n == 0:
        return fitz.Matrix(1, 0, 0, 1, 0, 0)
    # 1) 강인한 씨앗: 좌표축별 평행이동의 중앙값(median)은 소수 이상치에 안 흔들린다.
    dxs = sorted(dp[0] - sp[0] for sp, dp in pairs)
    dys = sorted(dp[1] - sp[1] for sp, dp in pairs)
    tdx = dxs[n // 2]
    tdy = dys[n // 2]
    if n < 6:
        # 표본이 적으면 절사가 오히려 불안정 → 평행이동만(강인한 중앙값).
        return fitz.Matrix(1, 0, 0, 1, tdx, tdy)
    # 2) 순수 평행이동 기준 잔차로 정렬해 하위 70%만 내부값으로 채택(상위 30%
    #    = 오매칭 또는 국소 이동 설비는 전역 배율 추정에서 제외).
    scored = sorted(
        pairs,
        key=lambda p: math.hypot(p[1][0] - (p[0][0] + tdx),
                                 p[1][1] - (p[0][1] + tdy)),
    )
    keep = max(4, int(len(scored) * 0.70))
    inliers = scored[:keep]
    # 3) 내부값으로만 배율+평행이동 재적합.
    s, sx, sy, dx, dy = _fit_scale_translate(inliers)
    if s is None or not math.isfinite(s) or abs(s - 1.0) > 0.05:
        # 같은 템플릿 시트에서 ±5% 초과 배율은 오매칭 잔재 → 순수 평행이동.
        return fitz.Matrix(1, 0, 0, 1, tdx, tdy)
    return fitz.Matrix(s, 0, 0, s, dx - s * sx, dy - s * sy)


# 국소 보정(IDW·Tag 소속)에서 마크업 하나가 끌려갈 수 있는 최대 평행이동(pt).
# Train copy 사이 같은 설비의 실제 재배치는 길어야 수십~150pt 수준이다. 이보다
# 큰 이동은 오매칭된 기준점이 끌어당기는 것이므로 상한으로 잘라 폭주를 막는다.
LOCAL_MAX_OFFSET = 200.0


def _inlier_pairs(pairs, base_matrix):
    """전역 변환(base_matrix) 적용 후 잔차가 비정상적으로 큰 기준점 쌍(오매칭
    또는 페이지를 가로지르는 이동)을 제외한 목록을 돌려준다. 이런 쌍이 IDW나
    Tag 소속 보정에 남아 있으면 주변 마크업을 수백 pt씩 끌어당겨 폭주시킨다.
    강인한 기준: 잔차가 (중앙값 + 3·MAD)와 절대 하한(LOCAL_MAX_OFFSET) 중
    큰 값을 넘으면 이상치로 본다. 내부값이 너무 적게 남으면 원본을 유지한다."""
    n = len(pairs)
    if n < 6:
        return pairs
    res = []
    for sp, dp in pairs:
        pred = fitz.Point(sp[0], sp[1]) * base_matrix
        res.append(math.hypot(dp[0] - pred.x, dp[1] - pred.y))
    sres = sorted(res)
    med = sres[n // 2]
    devs = sorted(abs(r - med) for r in res)
    mad = devs[n // 2]
    thresh = max(med + 3.0 * (mad if mad > 1e-6 else 1.0), LOCAL_MAX_OFFSET)
    inliers = [p for p, r in zip(pairs, res) if r <= thresh]
    if len(inliers) < max(4, n // 2):
        return pairs
    return inliers


def _clamp_offset(ox, oy, limit=LOCAL_MAX_OFFSET):
    """국소 보정 평행이동의 크기를 limit 이내로 자른다(방향은 보존)."""
    mag = math.hypot(ox, oy)
    if mag <= limit or mag < 1e-9:
        return ox, oy
    k = limit / mag
    return ox * k, oy * k


def _idw_offset(center, pairs, base_matrix, power=2.0):
    """전역 변환(base_matrix)을 적용한 뒤 남는 국소 위치 오차를, center(소스 좌표)
    주변 기준점들의 '실제 어긋남(잔차)'을 거리 역가중(IDW)으로 보간해 평행이동
    보정량 (ox, oy)로 돌려준다. 마크업이 기준점에 가까울수록 그 기준점이 가리키는
    정확한 위치로 끌려간다. 평행이동만이라 마크업 모양/기울기는 변하지 않고,
    서로 가까운 멀티파트(구름+콜아웃)는 거의 같은 양만큼 이동해 어긋나지 않는다."""
    cx, cy = center[0], center[1]
    total_w = 0.0
    ox = oy = 0.0
    for sp, dp in pairs:
        pred = fitz.Point(sp[0], sp[1]) * base_matrix  # 전역 변환 후 예측 위치
        rx = dp[0] - pred.x          # 잔차(목표 - 예측), 대상 좌표
        ry = dp[1] - pred.y
        d2 = (cx - sp[0]) ** 2 + (cy - sp[1]) ** 2
        if d2 < 1e-6:
            return (rx, ry)          # 기준점 위에 정확히 있는 마크업
        w = 1.0 / (d2 ** (power / 2.0))
        ox += w * rx
        oy += w * ry
        total_w += w
    if total_w <= 0:
        return (0.0, 0.0)
    return (ox / total_w, oy / total_w)


# 클러스터(구름+지시선+콜아웃 묶음)가 '어떤 계장 Tag를 설명하는 마크업인지'를
# 찾기 위한 최대 거리. 도면을 그리는 사람마다 CAD 배치가 달라져도(=도면이
# 틀어져도) 마크업이 가리키는 실제 설비(Tag)는 동일하므로, 그 Tag 하나의
# 정확한 이동량을 클러스터 전체에 우선 적용하면 평균치로 뭉개지는 일반 IDW
# 보정보다 훨씬 정확하다. 너무 멀리 있는 Tag까지 엮이면 오매칭이 되므로
# 적당한 거리 안에 있을 때만 채택한다.
TAG_ASSOC_MAX_DIST = 200.0  # pt


def _cluster_tag_anchor(src_center, named_pairs):
    """클러스터의 Source 좌표 중심(src_center) 근처에서 가장 가까운 named anchor
    Tag를 찾아 (tag_name, src_pt, dst_pt)를 돌려준다. 적당한 거리 안에 없으면 None."""
    if not named_pairs:
        return None
    cx, cy = src_center
    best = None
    best_d = None
    for name, sp, dp in named_pairs:
        d = math.hypot(cx - sp[0], cy - sp[1])
        if best_d is None or d < best_d:
            best_d = d
            best = (name, sp, dp)
    if best is not None and best_d <= TAG_ASSOC_MAX_DIST:
        return best
    return None


# ── CAD 형상 스냅(Target 도면의 실제 원/선에 마크업을 달라붙임) ───────────────
# 기준점(IDW) 보정으로도 못 잡는 미세 어긋남을, Target PDF에 실제로 그려진
# 도형(계장 심볼 원 / 배관선)에 마크업을 '가까울 때만' 달라붙여 해결한다.
SNAP_ENABLED = True       # 형상 스냅 전체 토글
SNAP_LINE_DIST = 14.0     # 선에 이 거리 안이면 선 위로 스냅(pt)
SNAP_MAX_MOVE = 60.0      # 스냅 이동량이 이보다 크면 오스냅으로 보고 무시(pt)
SNAP_CLOUD_SIZE = 70.0    # 이보다 큰 마크업(구름 등)은 원중심 스냅 판정에서 제외(pt)
SNAP_LINE_DIST_FORCED = 55.0  # 라인 강제 대상(X·Spectacle blind 등)의 넓은 탐색 거리(pt)
SNAP_MAX_MOVE_FORCED = 90.0   # 라인 강제 대상이 선에 붙기 위해 허용하는 최대 이동(pt)


def _is_x_mark_annot(annot) -> bool:
    """Ink로 그려진 'X' 표시(대각선 두 개가 교차)인지 판별한다. Symbol and
    Legend의 INSTR. PIPING NOTE 표기상 X는 Capillary/계기배관이 공정선에
    연결되는 지점을 뜻하므로, 항상 라인 스냅 대상이어야 하고 원중심
    (Tag 버블) 스냅 대상에서는 제외해야 한다."""
    try:
        if annot.type[1] != "Ink":
            return False
        strokes = annot.vertices
        if not strokes or len(strokes) != 2:
            return False
        angles = []
        for stroke in strokes:
            if len(stroke) != 2:
                return False
            (x0, y0), (x1, y1) = stroke
            if math.hypot(x1 - x0, y1 - y0) < 1e-6:
                return False
            angles.append(math.atan2(y1 - y0, x1 - x0))
        diff = abs(angles[0] - angles[1]) % math.pi
        diff = min(diff, math.pi - diff)
        return math.radians(40) < diff < math.radians(140)
    except Exception:
        return False


def _extract_target_geometry(page):
    """Target 페이지의 벡터 도형을 추출해 원(계장 심볼)·수평선·수직선 목록으로
    돌려준다(페이지에 캐시). 좌표는 fitz 좌표계."""
    cache = getattr(page, "_snap_geom_cache", None)
    if cache is not None:
        return cache
    circles = []   # (cx, cy, r)
    hlines = []    # (x0, x1, y)
    vlines = []    # (x, y0, y1)
    try:
        drawings = page.get_drawings()
    except Exception:
        drawings = []
    for d in drawings:
        rect = d.get("rect")
        items = d.get("items", []) or []
        n_curve = sum(1 for it in items if it and it[0] == "c")
        # 원/타원: 곡선 위주 + bbox가 거의 정사각 + 심볼 크기 범위
        if rect is not None and n_curve >= 2:
            w, h = rect.width, rect.height
            if w > 4 and h > 4 and 0.65 < (w / h) < 1.55 and max(w, h) < 130:
                circles.append(((rect.x0 + rect.x1) / 2.0,
                                (rect.y0 + rect.y1) / 2.0, max(w, h) / 2.0))
        # 직선 세그먼트
        for it in items:
            if not it or it[0] != "l":
                continue
            p1, p2 = it[1], it[2]
            if abs(p1.y - p2.y) < 0.6 and abs(p1.x - p2.x) > 8:
                hlines.append((min(p1.x, p2.x), max(p1.x, p2.x), (p1.y + p2.y) / 2.0))
            elif abs(p1.x - p2.x) < 0.6 and abs(p1.y - p2.y) > 8:
                vlines.append(((p1.x + p2.x) / 2.0, min(p1.y, p2.y), max(p1.y, p2.y)))
    cache = {"circles": circles, "hlines": hlines, "vlines": vlines}
    try:
        page._snap_geom_cache = cache
    except Exception:
        pass
    return cache


# 범례("INSTRUMENT IDENTIFICATION/계기 식별")의 표준 표기: 계기 Tag는 원을
# 가로지르는 선으로 위/아래 두 줄로 나뉘어, 위에는 기능문자(PDI/LT/FT/TT 등),
# 아래에는 루프번호(0152, 0151A 등)가 적힌다. PDF 텍스트 추출에서는 이 두 줄이
# 보통 별개 단어로 떨어져 나오기 때문에(예: "PDI", "0152"), 하이픈 붙은 한 단어만
# 잡는 _extract_generic_tags 로는 거의 인식이 안 된다. 원 안의 텍스트를 위/아래로
# 나눠 합쳐 "PDI-0152" 형태로 재구성하면, 구름(클라우드) 마크업이 감싸는 계기를
# 정확히 식별할 수 있어 클러스터→Tag 소속 보정(_cluster_tag_anchor)의 정확도가
# 크게 올라간다.
_BUBBLE_FUNC_RE = re.compile(r'^[A-Za-z]{1,6}$')
_BUBBLE_LOOP_RE = re.compile(r'^\d{2,6}[A-Za-z]?$')


def _extract_instrument_bubbles(page):
    """페이지의 계기 버블(원+가로 분할선+위/아래 텍스트)을 찾아 'FUNC-LOOP'
    Tag 문자열별 중심점을 반환한다. 같은 Tag가 여러 번 나오면 모호하므로 제외."""
    geom = _extract_target_geometry(page)
    circles = geom["circles"]
    if not circles:
        return {}
    try:
        words = page.get_text("words")
    except Exception:
        return {}
    tag_map = {}
    for ccx, ccy, r in circles:
        if r < 5:
            continue
        upper, lower = [], []
        for w in words:
            x0, y0, x1, y1, text = w[0], w[1], w[2], w[3], w[4]
            wx, wy = (x0 + x1) / 2.0, (y0 + y1) / 2.0
            if math.hypot(wx - ccx, wy - ccy) > r * 1.05:
                continue
            (upper if wy < ccy else lower).append((x0, text))
        if not upper or not lower:
            continue
        up_text = "".join(t for _, t in sorted(upper))
        low_text = "".join(t for _, t in sorted(lower))
        if not _BUBBLE_FUNC_RE.match(up_text) or not _BUBBLE_LOOP_RE.match(low_text):
            continue
        tag = f"{up_text}-{low_text}"
        if tag in tag_map:
            tag_map[tag] = None  # 중복 발견 → 모호함 표시
        else:
            tag_map[tag] = (ccx, ccy)
    return {k: v for k, v in tag_map.items() if v is not None}


def _cluster_indices(rects, gap=12.0):
    """transform된 마크업 rect들을 근접(겹치거나 gap 이내)끼리 묶어 클러스터
    id 리스트를 돌려준다(union-find). 구름+지시선+콜아웃처럼 붙어있는 멀티파트가
    한 덩어리로 묶여, 스냅을 클러스터 단위로 똑같이 적용해 안 어긋나게 한다."""
    n = len(rects)
    parent = list(range(n))

    def find(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    def union(i, j):
        ri, rj = find(i), find(j)
        if ri != rj:
            parent[rj] = ri

    def near(a, b):
        return not (a.x0 - gap > b.x1 or b.x0 - gap > a.x1 or
                    a.y0 - gap > b.y1 or b.y0 - gap > a.y1)

    for i in range(n):
        for j in range(i + 1, n):
            if near(rects[i], rects[j]):
                union(i, j)
    return [find(i) for i in range(n)]


def _snap_offset_for_cluster(centers, rects, geom, force_line_list=None):
    """클러스터에 속한 마크업들의 중심/rect를 보고, Target 도형(원/선)에
    '가까울 때만' 달라붙도록 클러스터 전체에 적용할 평행이동 (dx,dy)를 돌려준다.
    스냅 대상이 없거나 이동량이 과하면 (0,0).
    force_line_list: 멤버별 '반드시 라인 위에 있어야 하는' 마크업 여부(같은 순서).
    X(Capillary 연결점), Spectacle blind/Reducer 등 소스에서 공정선에 얹혀 있던
    마크업이 해당된다. 이들은 원중심 스냅에서 제외하고, 라인이 더 멀리(타깃에서
    이동) 있어도 붙도록 더 넓은 탐색 거리로 강제 라인 스냅한다."""
    if force_line_list is None:
        force_line_list = [False] * len(centers)
    best = None
    best_d = None
    # 1) 계장 Tag → 심볼 원 중심: 멤버 중심이 어떤 원 안(반경*1.3)에 들면 그 중심으로.
    # 구름(클라우드)처럼 큰 마크업은 자기 자신이 심볼 영역을 통째로 덮고 있어
    # 중심이 우연히 원 안에 들어가는 경우가 많으므로 판정에서 제외한다(작은
    # Tag번호 박스/계측기 버블만 이 규칙의 대상으로 삼는다). 라인 강제 대상
    # (X·Spectacle blind 등 공정선에 얹힌 것)도 원중심이 아니라 선에 붙어야 한다.
    for (cx, cy), rc, force in zip(centers, rects, force_line_list):
        if force or max(rc.width, rc.height) > SNAP_CLOUD_SIZE:
            continue
        for ccx, ccy, r in geom["circles"]:
            d = math.hypot(cx - ccx, cy - ccy)
            if d < r * 1.3 and (best_d is None or d < best_d):
                best_d = d
                best = (ccx - cx, ccy - cy)
    if best is not None:
        if math.hypot(*best) <= SNAP_MAX_MOVE:
            return best, "원중심"
        return (0.0, 0.0), None
    # 2) Reducer/MIN/Capillary X/Spectacle blind → 가장 가까운 선 위로 스냅.
    # 작은 마크업과 라인 강제 대상만(강제 대상은 크기 무관, 더 넓은 탐색 거리/
    # 더 큰 이동 허용 — 타깃에서 라인이 멀리 옮겨가도 다시 붙게).
    best = None
    best_d = None
    for (cx, cy), rc, force in zip(centers, rects, force_line_list):
        size = max(rc.width, rc.height)
        if size > 160 and not force:
            continue  # 너무 큰 마크업은 선 스냅 제외(큰 구름 등)
        line_dist = SNAP_LINE_DIST_FORCED if force else SNAP_LINE_DIST
        for x0, x1, y in geom["hlines"]:
            if x0 - 6 <= cx <= x1 + 6:
                d = abs(cy - y)
                if d < line_dist and (best_d is None or d < best_d):
                    best_d = d
                    best = (0.0, y - cy)
        for x, y0, y1 in geom["vlines"]:
            if y0 - 6 <= cy <= y1 + 6:
                d = abs(cx - x)
                if d < line_dist and (best_d is None or d < best_d):
                    best_d = d
                    best = (x - cx, 0.0)
    if best is not None:
        # 강제 라인 대상은 더 큰 이동까지 허용(허공에 남느니 선에 붙는 게 맞다).
        any_force = any(force_line_list)
        cap = SNAP_MAX_MOVE_FORCED if any_force else SNAP_MAX_MOVE
        if math.hypot(*best) <= cap:
            return best, "라인"
    return (0.0, 0.0), None


# 리듀서 사이즈 표기 라벨: "150x50","8X50" 등 (숫자+x/×+숫자가 한 토큰). 그 근처에
# 따로 찍힌 숫자 마크업("50")은 리듀서 출구 사이즈를 강조한 것이므로, 도면이
# 틀어져도 이 라벨 바로 옆을 따라가야 한다(허공에 뜨면 안 됨). 같은 위치라도
# Train copy마다 실제 배관 사이즈 값 자체는 다를 수 있으므로(예: 소스 "150x50"
# ↔ 타깃 "150x40"), 텍스트 일치가 아니라 '위치상 가장 가까운 라벨'로 대응시킨다 —
# 값이 달라도 같은 리듀서 위치를 가리키는 라벨이면 맞는 대응이다.
_REDUCER_SIZE_RE = re.compile(r'^\d{1,4}\s*[xX×]\s*\d{1,4}$')  # "150x50","8X50" 등
_NUM_ONLY_RE = re.compile(r'^\d{1,4}$')   # 리듀서 사이즈 마크업은 순수 숫자("50")
REDUCER_ADJ_DIST = 60.0    # 소스에서 숫자 마크업이 리듀서 라벨에 '붙어 있다'고 볼 거리
REDUCER_MATCH_DIST = 160.0  # 타깃에서 대응 리듀서 라벨을 찾는 최대 반경(값은 달라도 됨)


def _extract_reducer_tokens(page):
    """페이지에서 리듀서 사이즈 표기 토큰('80X','80×' 등)의 중심점 목록
    [(cx, cy, text), ...]을 돌려준다(페이지에 캐시)."""
    cache = getattr(page, "_reducer_tok_cache", None)
    if cache is not None:
        return cache
    out = []
    try:
        for w in page.get_text("words"):
            t = (w[4] or "").strip()
            if _REDUCER_SIZE_RE.match(t):
                out.append(((w[0] + w[2]) / 2.0, (w[1] + w[3]) / 2.0, t))
    except Exception:
        pass
    try:
        page._reducer_tok_cache = out
    except Exception:
        pass
    return out


def _nearest_token(pt, tokens, max_dist):
    """tokens([(cx,cy,text),...]) 중 pt에 가장 가까운 토큰을 max_dist 이내에서
    찾아 (cx, cy)를 돌려준다. 없으면 None."""
    best = None
    best_d = None
    for tx, ty, _ in tokens:
        d = math.hypot(pt[0] - tx, pt[1] - ty)
        if d <= max_dist and (best_d is None or d < best_d):
            best_d = d
            best = (tx, ty)
    return best


def _is_on_line(pt, geom, dist=6.0):
    """점 pt가 geom의 수평/수직선 위(거리 dist 이내)에 얹혀 있는지 판정한다.
    소스에서 라인에 붙어 있던 마크업(X·Spectacle blind·Reducer 등)을 찾아
    타깃에서도 라인에 강제로 다시 붙이기 위한 1차 분류용."""
    cx, cy = pt
    for x0, x1, y in geom["hlines"]:
        if x0 - 6 <= cx <= x1 + 6 and abs(cy - y) <= dist:
            return True
    for x, y0, y1 in geom["vlines"]:
        if y0 - 6 <= cy <= y1 + 6 and abs(cx - x) <= dist:
            return True
    return False


def _is_finite_point(pt) -> bool:
    return math.isfinite(pt[0]) and math.isfinite(pt[1]) and \
        abs(pt[0]) < 1_000_000 and abs(pt[1]) < 1_000_000


def _is_finite_rect(rect) -> bool:
    return all(math.isfinite(v) for v in (rect.x0, rect.y0, rect.x1, rect.y1)) and \
        abs(rect.x0) < 1_000_000 and abs(rect.y0) < 1_000_000 and \
        abs(rect.x1) < 1_000_000 and abs(rect.y1) < 1_000_000 and \
        rect.width > 0.01 and rect.height > 0.01


# PyMuPDF로 직접 재생성 가능한 마크업 타입(geometry 기반)
_SUPPORTED_TRANSFORM_TYPES = {
    "Square", "Circle", "Line", "PolyLine", "Polygon",
    "FreeText", "Highlight", "Ink", "StrikeOut", "Underline", "Squiggly",
}

# FreeText에서 쓰이는 일반 글꼴 이름을 PyMuPDF의 기본 14종 글꼴로 매핑.
# (임베드되지 않은 글꼴은 정확히 재현할 수 없어 가장 비슷한 기본 글꼴로 대체)
_FONT_NAME_MAP = {
    "helvetica": "helv", "arial": "helv", "arial narrow": "helv",
    "arial black": "helv", "calibri": "helv", "verdana": "helv",
    "times new roman": "tiro", "times": "tiro", "georgia": "tiro",
    "courier new": "cour", "courier": "cour", "consolas": "cour",
}


def _map_freetext_fontname(name) -> str:
    if not name:
        return "helv"
    return _FONT_NAME_MAP.get(name.strip().lower(), "helv")


def _parse_freetext_style(doc, xref):
    """FreeText 마크업의 /DS(Default Style) 문자열에서 글꼴명/크기/글자색을,
    /Q에서 정렬을 읽어온다. 값이 없으면 합리적인 기본값을 반환."""
    fontsize, fontname, text_color = 11.0, "helv", (0, 0, 0)
    try:
        kind, ds = doc.xref_get_key(xref, "DS")
    except Exception:
        kind, ds = None, None
    if kind == "string" and ds:
        m = re.search(r'font:\s*([\d.]+)pt\s+[\'"]?([^;\'"]+)[\'"]?', ds)
        if m:
            fontsize = float(m.group(1))
            fontname = _map_freetext_fontname(m.group(2))
        m2 = re.search(r'color:\s*#([0-9A-Fa-f]{6})', ds)
        if m2:
            hexcol = m2.group(1)
            text_color = tuple(int(hexcol[i:i + 2], 16) / 255 for i in (0, 2, 4))
    align = 0
    try:
        kind_q, q_val = doc.xref_get_key(xref, "Q")
        if kind_q == "int" and q_val:
            align = int(q_val)
    except Exception:
        pass
    return fontsize, fontname, text_color, align


def _annot_ap_normal_xref(doc, annot_xref):
    """annot의 AP(외형) Normal 스트림 xref를 반환. 없으면 None.
    /AP 값 자체가 인라인 딕셔너리(kind=="dict")가 아니라 별도 객체에 대한
    간접참조(kind=="xref")로 저장된 경우도 있다(FreeText 등에서 자주 보임) —
    이 경우를 놓치면 AP 클론이 실패해 폰트/색상이 깨지는 geometry 재생성
    방식으로 fallback되어 버린다."""
    try:
        kind, value = doc.xref_get_key(annot_xref, "AP")
    except Exception:
        return None
    if not value:
        return None
    if kind == "xref":
        m_ref = re.match(r'(\d+)\s+0\s+R', value.strip())
        if not m_ref:
            return None
        try:
            value = doc.xref_object(int(m_ref.group(1)), compressed=False)
        except Exception:
            return None
    elif kind != "dict":
        return None
    m = re.search(r'/N\s+(\d+)\s+0\s+R', value)
    return int(m.group(1)) if m else None


def _parse_pdf_floats(s):
    return [float(x) for x in re.findall(r'-?\d+(?:\.\d+)?', s)]


def _pdf_escape_text(text) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_text_string(text) -> str:
    """PDF 텍스트 문자열 리터럴을 만든다. 비ASCII(한글/러시아어 등)가 있으면
    UTF-16BE+BOM의 16진 문자열로 인코딩해야 한다 — 단순 리터럴 (...)로 쓰면
    PDFDocEncoding/Latin-1로 오인되어 깨진 글자(예: 모지바케)가 된다."""
    if all(ord(ch) < 128 for ch in text):
        return f"({_pdf_escape_text(text)})"
    raw = text.encode("utf-16-be")
    return f"<FEFF{raw.hex()}>"


def _graft_pdf_object(src_doc, dst_doc, src_xref, xref_map):
    """src_doc의 객체(src_xref)와 그 객체가 간접참조하는 모든 하위 객체를
    dst_doc로 깊은 복사한다. 외형 스트림(Form XObject)을 복사할 때 그 안의
    /Resources(글꼴 등)까지 같이 가져오기 위함 — 스트림 바이트만 복사하면
    글꼴 자원이 빠져 Helvetica 기본값으로 깨진다.
    xref_map: {원본xref: 새xref} (순환참조 방지 및 중복복사 방지)."""
    if src_xref in xref_map:
        return xref_map[src_xref]
    new_xref = dst_doc.get_new_xref()
    xref_map[src_xref] = new_xref  # 순환참조 대비 먼저 등록
    try:
        obj_str = src_doc.xref_object(src_xref, compressed=False)
    except Exception:
        return new_xref

    def repl(m):
        ref = int(m.group(1))
        return f"{_graft_pdf_object(src_doc, dst_doc, ref, xref_map)} 0 R"

    new_str = re.sub(r'\b(\d+)\s+0\s+R\b', repl, obj_str)

    is_stream = False
    try:
        is_stream = src_doc.xref_is_stream(src_xref)
    except Exception:
        is_stream = False

    if is_stream:
        data = None
        try:
            data = src_doc.xref_stream(src_xref)  # 압축 해제된 바이트
        except Exception:
            data = None
        if data is not None:
            # /Filter·/Length는 update_stream(compress=True)가 다시 채우므로 제거
            new_str = re.sub(r'/Filter\s*(?:/[A-Za-z0-9]+|\[[^\]]*\])', '', new_str)
            new_str = re.sub(r'/Length\s+\d+', '', new_str)
            dst_doc.update_object(new_xref, new_str)
            dst_doc.update_stream(new_xref, data, compress=True)
        else:
            # 압축 해제 실패(예: 이미지) → 원본 raw 바이트 그대로 복사
            try:
                raw = src_doc.xref_stream_raw(src_xref)
                dst_doc.update_object(new_xref, new_str)
                dst_doc.update_stream(new_xref, raw, compress=False)
            except Exception:
                dst_doc.update_object(new_xref, new_str)
    else:
        dst_doc.update_object(new_xref, new_str)
    return new_xref


# 원본 마크업에서 새 마크업으로 그대로 옮겨야 하는 스타일/속성 키들.
# 특히 /DA·/DS는 Bluebeam이 글꼴/크기/색상 속성을 읽어오는 곳이라, 이걸
# 복사하지 않으면 속성창에 Helvetica 같은 기본값이 표시된다(외형 자체는
# AP로 보존되더라도 별개).
_ANNOT_STYLE_KEYS = (
    "DA", "DS", "Q", "RC", "C", "IC", "IT", "CL", "BS", "BE",
    "RD", "CA", "Rotate", "Name", "NM", "Border", "BorderStyle",
)


def _clone_annot_with_appearance(src_annot, dst_page, matrix) -> bool:
    """원본 마크업의 실제 외형(AP, appearance stream)을 그대로 재사용해
    위치/회전/스케일만 적용한다. Bluebeam의 구름(cloud) 테두리, 글꼴/색상 등
    PyMuPDF의 geometry 재생성 방식으로는 재현이 안 되는 모든 커스텀 외형을
    원본 그대로 보존하기 위해, 새로 그리지 않고 원본 그림 자체를 transform한다."""
    src_doc = src_annot.parent.parent
    dst_doc = dst_page.parent
    ap_xref = _annot_ap_normal_xref(src_doc, src_annot.xref)
    if ap_xref is None:
        return False
    try:
        stream = src_doc.xref_stream(ap_xref)
        form_obj = src_doc.xref_object(ap_xref, compressed=False)
    except Exception:
        return False
    if not stream or not form_obj:
        return False

    bbox_m = re.search(r'/BBox\s*\[([^\]]+)\]', form_obj)
    if not bbox_m:
        return False
    bx = _parse_pdf_floats(bbox_m.group(1))
    if len(bx) != 4:
        return False
    bbox = fitz.Rect(bx[0], bx[1], bx[2], bx[3])

    mat_m = re.search(r'/Matrix\s*\[([^\]]+)\]', form_obj)
    if mat_m:
        mv = _parse_pdf_floats(mat_m.group(1))
        old_matrix = fitz.Matrix(*mv) if len(mv) == 6 else fitz.Matrix(1, 0, 0, 1, 0, 0)
    else:
        old_matrix = fitz.Matrix(1, 0, 0, 1, 0, 0)

    # BBox는 Form 내부의 로컬 좌표일 뿐이며, 실제 페이지상의 절대 위치는
    # PDF 스펙(12.5.5)에 따라 (BBox를 Matrix로 변환한 뒤) 그 결과를 Annotation의
    # 실제 Rect에 맞춰주는 별도의 보정행렬(AA)에서 나온다. 이 AA를 빼고 BBox만
    # 가지고 새 위치를 계산하면 원본 Rect 정보가 통째로 사라져 마크업이 전부
    # 원점 근처로 쏠려버린다. 원본 Rect를 반영해 AA를 직접 구해 합성한다.
    #
    # 주의: 외형 스트림/BBox/Matrix/Rect는 모두 PDF 원좌표계(y가 위로 증가)에
    # 있다. PyMuPDF의 annot.rect는 y-down으로 변환된 값이라 좌표계가 달라
    # 섞으면 상하가 뒤집힌다. 여기서는 원본 /Rect를 PDF 원좌표로 직접 읽는다.
    try:
        _, rect_val = src_doc.xref_get_key(src_annot.xref, "Rect")
        rv = _parse_pdf_floats(rect_val or "")
    except Exception:
        rv = []
    if len(rv) != 4:
        return False
    orig_rect = fitz.Rect(min(rv[0], rv[2]), min(rv[1], rv[3]),
                          max(rv[0], rv[2]), max(rv[1], rv[3]))
    orig_corners = [
        fitz.Point(bbox.x0, bbox.y0) * old_matrix,
        fitz.Point(bbox.x1, bbox.y0) * old_matrix,
        fitz.Point(bbox.x1, bbox.y1) * old_matrix,
        fitz.Point(bbox.x0, bbox.y1) * old_matrix,
    ]
    tb_xs = [c.x for c in orig_corners]
    tb_ys = [c.y for c in orig_corners]
    tb = fitz.Rect(min(tb_xs), min(tb_ys), max(tb_xs), max(tb_ys))
    sx = orig_rect.width / tb.width if tb.width > 1e-6 else 1.0
    sy = orig_rect.height / tb.height if tb.height > 1e-6 else 1.0
    aa = fitz.Matrix(sx, 0, 0, sy,
                      orig_rect.x0 - tb.x0 * sx, orig_rect.y0 - tb.y0 * sy)

    # 위치 보정 matrix는 PyMuPDF의 fitz 좌표계(y-down, MediaBox 원점 보정 포함)에서
    # 구한 것이다. 외형(AP)은 PDF 원좌표계(y-up)에 있으므로 좌표계를 변환해야 한다.
    # 페이지 높이로 단순히 y를 뒤집는 방식은 MediaBox 원점이 (0,0)이 아니거나
    # 페이지가 회전(/Rotate)된 도면에서 어긋나 마크업이 화면 밖으로 날아간다.
    # PyMuPDF가 제공하는 정확한 페이지 변환행렬을 쓴다:
    #   transformation_matrix : PDF → fitz,  그 역행렬 : fitz → PDF
    t_src = src_annot.parent.transformation_matrix      # PDF → fitz (Source)
    t_dst_inv = ~dst_page.transformation_matrix         # fitz → PDF (Target)
    matrix_pdf = t_src * matrix * t_dst_inv

    combined = old_matrix * aa * matrix_pdf
    corners = [
        fitz.Point(bbox.x0, bbox.y0) * combined,
        fitz.Point(bbox.x1, bbox.y0) * combined,
        fitz.Point(bbox.x1, bbox.y1) * combined,
        fitz.Point(bbox.x0, bbox.y1) * combined,
    ]
    if not all(_is_finite_point((c.x, c.y)) for c in corners):
        return False
    xs = [c.x for c in corners]
    ys = [c.y for c in corners]
    new_rect = fitz.Rect(min(xs), min(ys), max(xs), max(ys))
    if not _is_finite_rect(new_rect):
        return False

    subtype = src_annot.type[1]
    opacity = src_annot.opacity if src_annot.opacity is not None else 1
    content = (src_annot.info or {}).get("content", "") or ""

    try:
        # 핵심: 주석을 /Type·/Subtype·/Rect·/AP 등 일부 키만 골라 새로 조립하면
        # Bluebeam이 마크업 인식에 쓰는 고유 키(/NM, /Subj, /BE 구름효과, /IT
        # intent, 그룹/Popup 참조, Bluebeam 전용 데이터 등)가 전부 빠져, 외형은
        # 멀쩡히 저장돼도(=PNG·Adobe엔 보임) Bluebeam 화면엔 안 보이게 된다.
        # → 원본 주석 객체를 통째로 깊은 복사(graft)한 뒤, 좌표 관련 키만 덮어쓴다.
        #
        # /P(부모 페이지)가 소스 페이지를 통째로 복사하지 않도록, xref_map에
        # 소스 페이지 → 대상 페이지 매핑을 미리 심어둔다.
        src_page = src_annot.parent
        xref_map = {src_page.xref: dst_page.xref}
        new_annot_xref = _graft_pdf_object(src_doc, dst_doc, src_annot.xref, xref_map)

        # graft가 외형(AP) Form XObject도 /Resources(글꼴 등)까지 같이 복사했다.
        # 그 새 AP의 /Matrix만 위치/회전/스케일이 반영된 값으로 덮어쓴다.
        new_ap_xref = xref_map.get(ap_xref)
        if new_ap_xref:
            dst_doc.xref_set_key(
                new_ap_xref, "Matrix",
                f"[{combined.a:.6f} {combined.b:.6f} {combined.c:.6f} "
                f"{combined.d:.6f} {combined.e:.4f} {combined.f:.4f}]"
            )

        # 위치(Rect)·부모 페이지(P)를 대상 좌표/페이지로 교체.
        dst_doc.xref_set_key(
            new_annot_xref, "Rect",
            f"[{new_rect.x0:.4f} {new_rect.y0:.4f} {new_rect.x1:.4f} {new_rect.y1:.4f}]"
        )
        dst_doc.xref_set_key(new_annot_xref, "P", f"{dst_page.xref} 0 R")

        # 표시 플래그(/F): Print(4) 켜고, Hidden(2)·NoView(32) 끔 — 화면에 보이게.
        try:
            _, fval = src_doc.xref_get_key(src_annot.xref, "F")
            f_int = int(re.search(r'-?\d+', fval).group()) if fval else 0
        except Exception:
            f_int = 0
        f_int = (f_int | 4) & ~2 & ~32
        dst_doc.xref_set_key(new_annot_xref, "F", str(f_int))

        # Adobe류 뷰어는 /AP만 있으면 그대로 그려주지만, Bluebeam은 자체 마크업
        # 엔진이 타입별 필수 기하 키(Line=/L, Polygon·PolyLine=/Vertices,
        # Ink=/InkList, Highlight 등=/QuadPoints, FreeText 말풍선선=/CL)를 직접
        # 읽는다. graft로 가져온 원본 키는 '소스 좌표'라 대상 좌표계
        # (matrix_pdf)로 변환해 덮어써야 위치가 맞는다.
        def _xform_flat_pairs(raw_str):
            nums = _parse_pdf_floats(raw_str)
            out = []
            for i in range(0, len(nums) - 1, 2):
                p = fitz.Point(nums[i], nums[i + 1]) * matrix_pdf
                out.append(p.x)
                out.append(p.y)
            return out

        _GEOM_KEYS = ("L", "Vertices", "QuadPoints", "CL")
        for gkey in _GEOM_KEYS:
            try:
                gkind, gval = src_doc.xref_get_key(src_annot.xref, gkey)
            except Exception:
                continue
            if gkind != "array" or not gval:
                continue
            xs = _xform_flat_pairs(gval)
            if not xs:
                continue
            arr_str = "[" + " ".join(f"{v:.4f}" for v in xs) + "]"
            try:
                dst_doc.xref_set_key(new_annot_xref, gkey, arr_str)
            except Exception:
                pass

        try:
            ikind, ival = src_doc.xref_get_key(src_annot.xref, "InkList")
        except Exception:
            ikind, ival = "null", None
        if ikind == "array" and ival:
            strokes_raw = re.findall(r'\[([^\[\]]*)\]', ival)
            new_strokes = []
            for sraw in strokes_raw:
                xs = _xform_flat_pairs(sraw)
                if xs:
                    new_strokes.append(
                        "[" + " ".join(f"{v:.4f}" for v in xs) + "]"
                    )
            if new_strokes:
                try:
                    dst_doc.xref_set_key(
                        new_annot_xref, "InkList",
                        "[" + " ".join(new_strokes) + "]"
                    )
                except Exception:
                    pass

        page_xref = dst_page.xref
        kind, annots_val = dst_doc.xref_get_key(page_xref, "Annots")
        if kind == "array":
            # 페이지가 /Annots를 직접 배열로 들고 있는 경우: 끝에 추가
            inner = (annots_val or "").strip()
            if inner.startswith("[") and inner.endswith("]"):
                inner = inner[1:-1]
            dst_doc.xref_set_key(
                page_xref, "Annots", f"[{inner} {new_annot_xref} 0 R]"
            )
        elif kind == "xref":
            # /Annots가 별도 배열 객체를 간접참조하는 경우: 그 객체 자체에
            # 추가해야 한다(페이지의 키를 덮어쓰면 이전에 추가된 마크업과
            # 원본 마크업까지 전부 사라진다).
            m = re.match(r'(\d+)\s+0\s+R', (annots_val or "").strip())
            if not m:
                return False
            arr_xref = int(m.group(1))
            arr_str = dst_doc.xref_object(arr_xref, compressed=False) or "[]"
            inner = arr_str.strip()
            if inner.startswith("[") and inner.endswith("]"):
                inner = inner[1:-1]
            dst_doc.update_object(arr_xref, f"[{inner} {new_annot_xref} 0 R]")
        else:
            dst_doc.xref_set_key(page_xref, "Annots", f"[{new_annot_xref} 0 R]")
    except Exception:
        return False

    return True


def _copy_annot_with_transform(src_annot, dst_page, matrix):
    """src_annot의 geometry를 matrix로 변환해 dst_page에 동일한 종류의 마크업을 생성.
    지원하지 않는 타입이면 False 반환(스킵)."""
    subtype = src_annot.type[1]
    colors = src_annot.colors or {}
    stroke = colors.get("stroke")
    fill = colors.get("fill")
    border_info = src_annot.border or {}
    width = border_info.get("width")
    width = 1 if width is None else width  # 0(테두리 없음)은 그대로 0 유지
    clouds = border_info.get("clouds", 0) or 0
    dashes = border_info.get("dashes") or None
    opacity = src_annot.opacity if src_annot.opacity is not None else 1

    def tp(pt):
        """튜플/Point를 fitz.Point로 변환 후 matrix 적용"""
        return fitz.Point(pt[0], pt[1]) * matrix

    new_annot = None

    try:
        if subtype in ("Square",):
            rect = src_annot.rect * matrix
            if not _is_finite_rect(rect):
                return False
            new_annot = dst_page.add_rect_annot(rect)
        elif subtype == "Circle":
            rect = src_annot.rect * matrix
            if not _is_finite_rect(rect):
                return False
            new_annot = dst_page.add_circle_annot(rect)
        elif subtype in ("Line",):
            verts = src_annot.vertices
            pts = [tp(p) for p in verts] if verts else None
            if not pts:
                v = src_annot.line
                pts = [tp(v[0]), tp(v[1])] if v else None
            if not pts or len(pts) < 2 or not all(_is_finite_point(p) for p in pts):
                return False
            new_annot = dst_page.add_line_annot(pts[0], pts[1])
        elif subtype in ("PolyLine", "Polygon"):
            verts = src_annot.vertices
            if not verts:
                return False
            pts = [tp(v) for v in verts]
            if not all(_is_finite_point(p) for p in pts):
                return False
            if subtype == "PolyLine":
                new_annot = dst_page.add_polyline_annot(pts)
            else:
                new_annot = dst_page.add_polygon_annot(pts)
        elif subtype == "Ink":
            try:
                strokes = src_annot.ink_list
            except Exception:
                return False
            if not strokes:
                return False
            transformed = [[tp(pt) for pt in stroke] for stroke in strokes]
            if not all(_is_finite_point(p) for stroke in transformed for p in stroke):
                return False
            new_annot = dst_page.add_ink_annot(transformed)
        elif subtype == "FreeText":
            rect = src_annot.rect * matrix
            if not _is_finite_rect(rect):
                return False
            text = (src_annot.info or {}).get("content", "") or ""
            ft_scale = math.hypot(matrix.a, matrix.b)
            fontsize, fontname, text_color, align = _parse_freetext_style(
                src_annot.parent.parent, src_annot.xref
            )
            new_annot = dst_page.add_freetext_annot(
                rect, text,
                fontsize=fontsize * ft_scale,
                fontname=fontname,
                text_color=text_color,
                fill_color=fill,
                border_color=stroke,
                align=align,
            )
        elif subtype in ("Highlight", "StrikeOut", "Underline", "Squiggly"):
            quads = src_annot.vertices
            if not quads or len(quads) < 4:
                return False
            # vertices: 4개씩 끊어서 quad 단위로 변환
            flat = [tp(pt) for pt in quads]
            if not all(_is_finite_point(p) for p in flat):
                return False
            if subtype == "Highlight":
                new_annot = dst_page.add_highlight_annot(flat)
            elif subtype == "StrikeOut":
                new_annot = dst_page.add_strikeout_annot(flat)
            elif subtype == "Underline":
                new_annot = dst_page.add_underline_annot(flat)
            else:
                new_annot = dst_page.add_squiggly_annot(flat)
        else:
            return False
    except Exception:
        return False

    if new_annot is None:
        return False

    try:
        col_kwargs = {}
        if stroke:
            col_kwargs["stroke"] = stroke
        if fill:
            col_kwargs["fill"] = fill
        if col_kwargs:
            new_annot.set_colors(**col_kwargs)

        scale = math.hypot(matrix.a, matrix.b)
        border_kwargs = {"width": width * scale}
        if dashes:
            border_kwargs["dashes"] = dashes
        # Polygon/PolyLine은 원본 꼭짓점 자체가 이미 구름 모양이라(들쭐날쭐한 path를
        # 그대로 복사함) clouds를 다시 적용하면 구름 효과가 중복되어 모양이 깨진다.
        # Square/Circle은 단순 사각형/원으로만 재생성되므로 clouds가 필요하다.
        if clouds and subtype in ("Square", "Circle"):
            border_kwargs["clouds"] = clouds
        try:
            new_annot.set_border(**border_kwargs)
        except Exception:
            new_annot.set_border(width=width * scale)

        new_annot.set_opacity(opacity)
        new_annot.update()
    except Exception:
        pass

    return True


def copy_markups_with_position_correction(src_path, dst_path, out_path, log_fn=None):
    """도면 레이아웃이 다른 경우: Bluebeam 자동화 없이 PyMuPDF로 직접
    마크업 좌표를 보정해서 Target(→Output)에 복사한다.
    Source/Target에 마젠타색 기준점 마크업을 직접 찍어두면 그것을 우선 사용하고,
    없으면 배관선번호/계측기 Tag/Symbol/도면 모서리를 자동으로 찾아 매칭한다.
    반환: (copied, skipped) 마크업 개수"""
    def log(msg):
        if log_fn:
            log_fn(msg)

    if fitz is None:
        raise RuntimeError("PyMuPDF(fitz)가 설치되어 있지 않습니다. 'pip install PyMuPDF' 필요")

    src_doc = fitz.open(src_path)
    dst_doc = fitz.open(dst_path)

    match_result = _find_tag_matches(src_doc, dst_doc, log_fn=log)
    if not match_result:
        src_doc.close()
        dst_doc.close()
        raise RuntimeError(
            "위치 보정 실패: Source/Target에서 Train 번호만 다른 동일 Tag를 "
            "2개 이상 찾지 못했습니다."
        )

    src_page_idx, dst_page_idx, pairs, skip_xrefs, named_pairs = match_result

    # 전역 변환은 '회전 없는' 균일 스케일+평행이동만 쓴다(마크업이 통째로 기울지
    # 않도록). 남는 국소 위치 오차는 마크업마다 주변 기준점의 어긋남을 거리가중
    # 보간(IDW)한 '평행이동'으로 따로 보정한다 — 평행이동만이라 모양/기울기는
    # 안 변하고, 가까운 멀티파트(구름+콜아웃)는 거의 같은 양 이동해 안 어긋난다.
    base_matrix = _global_scale_translate_matrix(pairs)
    scale = base_matrix.a
    log(f"  [위치 보정] 기준점 {len(pairs)}개로 전역 변환 적용 "
        f"(회전 0° 고정, 배율 {scale:.4f}, 국소 IDW 보정 ON, "
        f"Source p{src_page_idx+1} → Target p{dst_page_idx+1})\n")

    src_page = src_doc[src_page_idx]
    dst_page = dst_doc[dst_page_idx]

    # 국소 보정용 기준점은 전역 변환 후 잔차가 비정상적으로 큰 오매칭을 제거한
    # 내부값만 쓴다(이상치가 IDW·Tag 소속 보정에 남으면 주변 마크업을 수백 pt씩
    # 끌어당겨 폭주시킨다 — 변위 진단의 최대 866/898pt 이동이 그 증상이었다).
    idw_pairs = _inlier_pairs(pairs, base_matrix)
    named_pairs = [(nm, sp_, dp_) for (nm, sp_, dp_) in named_pairs
                   if math.hypot(dp_[0] - (fitz.Point(sp_[0], sp_[1]) * base_matrix).x,
                                 dp_[1] - (fitz.Point(sp_[0], sp_[1]) * base_matrix).y)
                   <= LOCAL_MAX_OFFSET]
    n_drop = len(pairs) - len(idw_pairs)
    if n_drop > 0:
        log(f"  [위치 보정] 국소 보정에서 오매칭 추정 기준점 {n_drop}개 제외 "
            f"(잔차 과대 → 폭주 방지)\n")

    # 1차 패스: 마크업마다 전역+IDW 변환행렬과 transform된 rect/center를 미리 구한다.
    annots = [a for a in (src_page.annots() or []) if a.xref not in skip_xrefs]
    matrices = []   # 각 마크업의 base*IDW 행렬
    t_rects = []    # transform된 rect(fitz)
    t_centers = []  # transform된 중심
    s_centers = []  # Source 좌표 중심(클러스터→Tag 소속 판정에 사용)
    for a in annots:
        c = a.rect
        center = ((c.x0 + c.x1) / 2.0, (c.y0 + c.y1) / 2.0)
        ox, oy = _idw_offset(center, idw_pairs, base_matrix)
        ox, oy = _clamp_offset(ox, oy)
        m = base_matrix * fitz.Matrix(1, 0, 0, 1, ox, oy)
        tr = a.rect * m
        matrices.append(m)
        t_rects.append(tr)
        t_centers.append(((tr.x0 + tr.x1) / 2.0, (tr.y0 + tr.y1) / 2.0))
        s_centers.append(center)

    # 클러스터링: 구름+지시선+콜아웃처럼 붙어있는 멀티파트를 한 덩어리로 묶는다.
    # 이 클러스터 단위로 (1) Tag 소속 보정과 (2) CAD 형상 스냅을 둘 다 적용해
    # 묶음 안의 마크업들이 서로 어긋나지 않게 한다.
    from collections import defaultdict
    cl = _cluster_indices(t_rects)
    groups = defaultdict(list)
    for i, cid in enumerate(cl):
        groups[cid].append(i)

    # 클러스터→Tag 소속 보정: 그리는 사람마다 CAD 배치가 달라져 도면이 서로
    # 틀어지더라도, 마크업이 실제로 설명하는 설비(Tag)는 동일하다. 전체 평균
    # 변환(IDW)이 아니라 '이 클러스터가 소속된 Tag 하나'의 정확한 이동량을
    # 우선 적용해, 그 Tag가 도면 안에서 어디로 옮겨갔든 마크업이 따라가게 한다.
    n_tag_assoc = 0
    for members in groups.values():
        cx = sum(s_centers[i][0] for i in members) / len(members)
        cy = sum(s_centers[i][1] for i in members) / len(members)
        anchor = _cluster_tag_anchor((cx, cy), named_pairs)
        if anchor is None:
            continue
        tag_name, sp_, dp_ = anchor
        pred = fitz.Point(sp_[0], sp_[1]) * base_matrix
        tag_dx, tag_dy = dp_[0] - pred.x, dp_[1] - pred.y
        tag_dx, tag_dy = _clamp_offset(tag_dx, tag_dy)
        m_override = base_matrix * fitz.Matrix(1, 0, 0, 1, tag_dx, tag_dy)
        for i in members:
            matrices[i] = m_override
            tr = annots[i].rect * m_override
            t_rects[i] = tr
            t_centers[i] = ((tr.x0 + tr.x1) / 2.0, (tr.y0 + tr.y1) / 2.0)
        n_tag_assoc += 1
    if named_pairs:
        log(f"  [위치 보정] Tag 소속 보정: {n_tag_assoc}개 클러스터가 "
            f"가까운 Tag를 따라 이동\n")

    # 리듀서 사이즈 배치 보정: "50"처럼 숫자만 있는 마크업이 소스에서 리듀서
    # 사이즈 토큰("80X" 등) 바로 옆에 있으면, 이는 리듀서 출구 사이즈 표기다.
    # 타깃에서 '대응되는 리듀서 토큰'을 찾아 그 토큰을 따라가게 해(상대 위치
    # 보존), 도면이 틀어져도 허공이 아니라 리듀서 옆에 정확히 놓이게 한다.
    src_red = _extract_reducer_tokens(src_page)
    dst_red = _extract_reducer_tokens(dst_page)
    reducer_placed = set()
    n_reducer = 0
    if src_red and dst_red:
        for i, a in enumerate(annots):
            content = ((a.info or {}).get("content", "") or "").strip()
            if not _NUM_ONLY_RE.match(content):
                continue
            src_tok = _nearest_token(s_centers[i], src_red, REDUCER_ADJ_DIST)
            if src_tok is None:
                continue
            spred = fitz.Point(src_tok[0], src_tok[1]) * base_matrix
            dst_tok = _nearest_token((spred.x, spred.y), dst_red, REDUCER_MATCH_DIST)
            if dst_tok is None:
                continue
            # 마크업이 소스 토큰을 따라간 만큼(=토큰 변위) 평행이동.
            dxr, dyr = _clamp_offset(dst_tok[0] - spred.x, dst_tok[1] - spred.y)
            m_override = base_matrix * fitz.Matrix(1, 0, 0, 1, dxr, dyr)
            matrices[i] = m_override
            tr = annots[i].rect * m_override
            t_rects[i] = tr
            t_centers[i] = ((tr.x0 + tr.x1) / 2.0, (tr.y0 + tr.y1) / 2.0)
            reducer_placed.add(i)
            n_reducer += 1
        if n_reducer:
            log(f"  [위치 보정] 리듀서 사이즈 보정: 숫자 마크업 {n_reducer}개를 "
                f"타깃 리듀서 토큰 옆으로 이동\n")

    # 형상 스냅: 연결된 멀티파트가 안 어긋나도록 클러스터 단위로 같은 양만큼 스냅.
    snap_off = [(0.0, 0.0)] * len(annots)
    n_snap = {"원중심": 0, "라인": 0}
    if SNAP_ENABLED and annots:
        geom = _extract_target_geometry(dst_page)
        # 소스 형상도 뽑아, 각 마크업이 '소스에서 라인에 얹혀 있었는지' 판정한다.
        # 라인에 얹혀 있던 마크업(X·Spectacle blind·Reducer 등)은 타깃에서도
        # 반드시 라인 위에 있어야 하므로 강제 라인 스냅 대상으로 표시한다.
        src_geom = _extract_target_geometry(src_page)
        on_src_line = [_is_on_line(s_centers[i], src_geom) for i in range(len(annots))]
        log(f"  [스냅] Target 형상 추출: 원 {len(geom['circles'])}개 / "
            f"수평선 {len(geom['hlines'])}개 / 수직선 {len(geom['vlines'])}개\n")
        for members in groups.values():
            # 리듀서 사이즈로 이미 토큰 옆에 정확히 놓인 마크업은 스냅이 다시
            # 끌어가지 않도록 그 클러스터는 건너뛴다.
            if any(i in reducer_placed for i in members):
                continue
            ctr = [t_centers[i] for i in members]
            rcs = [t_rects[i] for i in members]
            force = [_is_x_mark_annot(annots[i]) or on_src_line[i] for i in members]
            (dx, dy), kind = _snap_offset_for_cluster(ctr, rcs, geom, force)
            if kind:
                n_snap[kind] += 1
                for i in members:
                    snap_off[i] = (dx, dy)
        log(f"  [스냅] 적용: 원중심 {n_snap['원중심']}개 클러스터 / "
            f"라인 {n_snap['라인']}개 클러스터\n")

    copied, skipped = 0, 0
    # 변위 진단: 단순 복사(원좌표 그대로) 대비 보정 후 마크업 중심이 실제로
    # 얼마나 움직였는지 픽셀로 측정한다. "보정해도 안 변한다"는 체감이
    # (a) 보정량이 실제 0에 가깝기 때문인지 (b) 보정은 되는데 화면에서 구분이
    # 안 되는지 구분하기 위함. 이동량이 큰 항목 몇 개도 같이 찍는다.
    moves = []
    for i, (sx, sy) in enumerate(snap_off):
        fcx = t_centers[i][0] + sx
        fcy = t_centers[i][1] + sy
        d = math.hypot(fcx - s_centers[i][0], fcy - s_centers[i][1])
        moves.append(d)
    if moves:
        mv_avg = sum(moves) / len(moves)
        mv_max = max(moves)
        n_moved = sum(1 for d in moves if d >= 2.0)
        log(f"  [위치 보정] 변위 진단: 마크업 {len(moves)}개 중 "
            f"{n_moved}개가 2pt 이상 이동 "
            f"(평균 {mv_avg:.1f}pt, 최대 {mv_max:.1f}pt). "
            f"이 값이 0에 가까우면 두 도면이 이미 정렬되어 보정 불필요 상태.\n")

    for a, m, (sx, sy) in zip(annots, matrices, snap_off):
        if sx or sy:
            m = m * fitz.Matrix(1, 0, 0, 1, sx, sy)
        ok = _clone_annot_with_appearance(a, dst_page, m)
        if not ok:
            ok = _copy_annot_with_transform(a, dst_page, m)
        if ok:
            copied += 1
        else:
            skipped += 1

    if skipped:
        log(f"  [위치 보정] ⚠ {skipped}개 마크업은 지원하지 않는 타입이라 스킵됨\n")
    log(f"  [위치 보정] {copied}개 마크업 좌표 보정하여 복사 완료\n")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    dst_doc.save(out_path)
    src_doc.close()
    dst_doc.close()

    # 저장 결과 자가진단: 출력 파일을 다시 열어 실제로 마크업이 페이지 안에
    # 들어있는지 확인한다(복사 개수는 양수인데 화면엔 안 보이는 경우, 마크업이
    # 페이지 밖으로 날아갔는지 / 아예 저장이 안 됐는지 구분하기 위함).
    try:
        chk = fitz.open(out_path)
        cp = chk[dst_page_idx]
        pr = cp.rect
        n_annot = 0
        n_onpage = 0
        bx0 = by0 = float("inf")
        bx1 = by1 = float("-inf")
        for a in cp.annots() or []:
            n_annot += 1
            r = a.rect
            bx0 = min(bx0, r.x0); by0 = min(by0, r.y0)
            bx1 = max(bx1, r.x1); by1 = max(by1, r.y1)
            if r.x1 > pr.x0 and r.x0 < pr.x1 and r.y1 > pr.y0 and r.y0 < pr.y1:
                n_onpage += 1
        # 외형(AP)이 실제로 '그려지는지' 검사: 각 주석의 외형을 단독 렌더링해
        # 빈(흰색/투명) 픽스맵이면 AP 자체가 깨진 것(=화면에 안 보이는 진짜 원인).
        n_empty_ap = 0
        n_render_ok = 0
        for a in cp.annots() or []:
            try:
                pm = a.get_pixmap()
            except Exception:
                n_empty_ap += 1
                continue
            # 모든 픽셀이 동일(=내용 없음)하면 빈 외형으로 본다. 큰 마크업
            # (구름 등)은 윗부분이 비어 있어 앞부분만 보면 멀쩡한데도 빈외형으로
            # 오판되므로, 버퍼 전체를 일정 간격(stride)으로 훑어 변화가 있는지 본다.
            try:
                samples = pm.samples
                if not samples:
                    n_empty_ap += 1
                else:
                    stride = max(1, len(samples) // 4096)
                    seen = set(samples[::stride])
                    if len(seen) <= 1:
                        n_empty_ap += 1
                    else:
                        n_render_ok += 1
            except Exception:
                n_empty_ap += 1
        # 출력된 각 마크업의 외형(AP) /Matrix에 실제로 회전이 들어갔는지 측정.
        # 도면이 안 돌아갔다면 모든 각도가 0°에 가까워야 한다. 0이 아니면 기울어짐의
        # 원인이 우리 변환(또는 원본 마크업 자체의 회전)임을 데이터로 확정한다.
        n_tilt = 0
        max_tilt = 0.0
        sum_tilt = 0.0
        n_meas = 0
        for a in cp.annots() or []:
            apx = _annot_ap_normal_xref(chk, a.xref)
            if apx is None:
                continue
            try:
                fo = chk.xref_object(apx, compressed=False)
            except Exception:
                continue
            mm = re.search(r'/Matrix\s*\[([^\]]+)\]', fo or "")
            if not mm:
                continue
            mv = _parse_pdf_floats(mm.group(1))
            if len(mv) != 6:
                continue
            ang = math.degrees(math.atan2(mv[1], mv[0]))
            # ±180/±90 부근(상하/좌우 반전 표현)은 회전 아님 → 0 기준 잔차만 본다
            ang = ((ang + 180) % 90) - 0  # 0~90 범위로 접되, 작은 잔차 판별용
            ang = min(abs(ang), abs(ang - 90))
            n_meas += 1
            sum_tilt += ang
            max_tilt = max(max_tilt, ang)
            if ang > 0.5:
                n_tilt += 1
        chk.close()
        if n_annot:
            log(f"  [자가진단] 출력 파일 주석 {n_annot}개 중 페이지 안 {n_onpage}개 "
                f"(페이지 {pr.width:.0f}x{pr.height:.0f}, "
                f"주석 영역 x[{bx0:.0f}~{bx1:.0f}] y[{by0:.0f}~{by1:.0f}])\n")
            log(f"  [자가진단] 외형 렌더링: 정상 {n_render_ok}개 / 빈외형 {n_empty_ap}개 "
                f"(빈외형이 많으면 AP 복사가 깨진 것)\n")
            if n_meas:
                log(f"  [자가진단] 마크업 기울기: 측정 {n_meas}개 중 0.5°↑ {n_tilt}개 "
                    f"(평균 {sum_tilt/n_meas:.2f}°, 최대 {max_tilt:.2f}°) "
                    f"— 0에 가까우면 우리 변환은 안 기울어진 것\n")
        else:
            log("  [자가진단] ⚠ 출력 파일에 주석이 하나도 저장되지 않았습니다\n")

        # 사람이 눈으로 확인할 수 있도록 페이지 전체를 PNG로도 저장한다.
        try:
            png_path = os.path.splitext(out_path)[0] + "_미리보기.png"
            chk2 = fitz.open(out_path)
            pix = chk2[dst_page_idx].get_pixmap(dpi=120, annots=True)
            pix.save(png_path)
            chk2.close()
            log(f"  [자가진단] 미리보기 PNG 저장: {png_path}\n")
        except Exception as e:
            log(f"  [자가진단] 미리보기 PNG 저장 실패: {e}\n")
    except Exception as e:
        log(f"  [자가진단] 확인 실패: {e}\n")

    return copied, skipped


def _position_correction_subprocess_main(src_path, dst_path, out_path, result_queue):
    """별도 프로세스에서 실행되는 진입점. PyMuPDF 네이티브 크래시가 나도
    이 프로세스만 죽고 메인 GUI 프로세스는 영향받지 않는다.
    내부 로그를 모아 메인 프로세스로 돌려보내(GUI 작업 로그에 표시) 디버깅을
    돕는다 — 그렇지 않으면 매칭/복사 과정이 전혀 안 보여 '왜 0개인지' 알 수 없다."""
    logs = []
    try:
        copied, skipped = copy_markups_with_position_correction(
            src_path, dst_path, out_path, log_fn=lambda m: logs.append(m)
        )
        result_queue.put(("ok", copied, skipped, logs))
    except Exception as e:
        logs.append(f"  💥 위치 보정 예외: {e}\n")
        result_queue.put(("error", str(e), logs))


def run_position_correction_isolated(src_path, dst_path, out_path,
                                     log_fn=None, timeout=90):
    """위치 보정을 별도 프로세스에서 실행해 네이티브 크래시로부터 메인 앱을 보호한다."""
    def log(msg):
        if log_fn:
            log_fn(msg)

    ctx = mp.get_context("spawn")
    result_queue = ctx.Queue()
    proc = ctx.Process(
        target=_position_correction_subprocess_main,
        args=(src_path, dst_path, out_path, result_queue),
    )
    proc.start()
    proc.join(timeout)

    if proc.is_alive():
        proc.terminate()
        proc.join()
        raise RuntimeError(f"위치 보정 작업이 {timeout}초 내에 끝나지 않아 중단했습니다.")

    try:
        result = result_queue.get_nowait()
    except Exception:
        exitcode = proc.exitcode
        raise RuntimeError(
            f"위치 보정 작업이 비정상 종료되었습니다 (exit code {exitcode}). "
            "PyMuPDF 내부 오류로 추정됩니다. Source/Target에 매칭 가능한 Tag가 있는지 확인하세요."
        )

    # 서브프로세스가 모은 내부 로그를 GUI 작업 로그로 그대로 출력
    for line in (result[-1] if isinstance(result[-1], list) else []):
        log(line)

    status = result[0]
    if status == "error":
        raise RuntimeError(result[1])
    return result[1], result[2]   # copied, skipped


def _open_and_copy_source(src: str, filter_settings=None, log_fn=None):
    """Source PDF를 열고 마크업을 클립보드에 복사한 뒤 닫는다.
    반환: (filter_kept, filter_removed, tmp_path_or_None)"""
    def log(msg):
        if log_fn:
            log_fn(msg)

    open_src = src
    filter_kept, filter_removed = None, None
    tmp_to_delete = None

    if filter_settings:
        log("  [필터] 마크업 필터 적용 중…\n")
        open_src, filter_kept, filter_removed = build_filtered_copy(
            src,
            filter_settings.get("color", ""),
            filter_settings.get("date_limit", ""),
            filter_settings.get("author", ""),
            log_fn=log,
        )
        if open_src != src:
            tmp_to_delete = open_src

    log(f"  [Source] 열기: {os.path.basename(src)}\n")
    open_pdf(open_src)
    fit_page()
    time.sleep(WAIT_SHORT)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(WAIT_SHORT)

    if filter_settings:
        log("  [필터] 병합 마크업 그룹 해제 중…\n")
        pyautogui.hotkey('ctrl', 'shift', 'g')
        time.sleep(WAIT_SHORT)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(WAIT_SHORT)

    pyautogui.hotkey('ctrl', 'c')
    time.sleep(WAIT_SHORT)

    log("  [Source] 마크업 복사 완료 → Source 닫기\n")
    close_pdf_discard()

    if tmp_to_delete:
        try:
            os.remove(tmp_to_delete)
        except OSError:
            pass

    return filter_kept, filter_removed


def _paste_to_target(dst: str, out: str, log_fn=None, stop_check=None):
    """클립보드 내용을 Target → Output에 붙여넣고 저장 후 해당 탭만 닫는다."""
    def log(msg):
        if log_fn:
            log_fn(msg)

    def check_stop():
        if stop_check and stop_check():
            raise StoppedError()

    check_stop()
    log(f"  [Target] Output 복사: {os.path.basename(out)}\n")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    shutil.copy2(dst, out)
    time.sleep(0.5)

    log(f"  [Target] 열기 → 마크업 붙여넣기 → 저장\n")
    open_pdf(out)
    fit_page()
    time.sleep(WAIT_SHORT)
    check_stop()

    pyautogui.hotkey('ctrl', 'shift', 'v')   # Paste in Place
    time.sleep(WAIT_PASTE)

    pyautogui.hotkey('ctrl', 's')            # Save
    time.sleep(2.5)
    pyautogui.press('enter')                 # 덮어쓰기 확인 팝업 대비
    time.sleep(1.0)

    # 탭만 닫기 (Bluebeam 앱은 유지) — 저장 완료 후라 저장 팝업 없음
    pyautogui.hotkey('ctrl', 'w')
    time.sleep(WAIT_SHORT)
    log("  ✓ 완료\n")


def process_pair(src: str, dst: str, out: str, log_fn=None, stop_check=None,
                 filter_settings=None):
    """단일 Source → 단일 Target 처리 (하위 호환용)"""
    def check_stop():
        if stop_check and stop_check():
            raise StoppedError()

    check_stop()
    filter_kept, filter_removed = _open_and_copy_source(src, filter_settings, log_fn)
    check_stop()
    _paste_to_target(dst, out, log_fn, stop_check)
    return {"filter_kept": filter_kept, "filter_removed": filter_removed}


def process_source_group(src: str, targets: list, log_fn=None, stop_check=None,
                         filter_settings=None):
    """Source 하나 → Target 여러 개 처리 (Source를 한 번만 열고 복사).
    targets: [(dst_path, out_path), ...]
    반환: [(filter_kept, filter_removed), ...] — 첫 Target 기준 필터 결과, 나머지는 동일"""
    def check_stop():
        if stop_check and stop_check():
            raise StoppedError()

    check_stop()
    filter_kept, filter_removed = _open_and_copy_source(src, filter_settings, log_fn)
    check_stop()

    results = []
    for dst, out in targets:
        _paste_to_target(dst, out, log_fn, stop_check)
        results.append((filter_kept, filter_removed))
        check_stop()
    return results


# ══════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(_APP_TITLE)
        self.resizable(True, True)
        self.minsize(900, 640)
        self._build_ui()

    # ── UI 최상위 ──────────────────────────────────────────

    def _build_ui(self):
        self.configure(bg="#1e1e2e")
        self._build_header()
        self._build_body()
        self._build_footer()

    def _build_header(self):
        hdr = tk.Frame(self, bg="#313244", pady=10)
        hdr.pack(fill="x")
        tk.Label(
            hdr, text="🔵  Bluebeam Markup Auto-Copy",
            font=("Segoe UI", 16, "bold"),
            fg="#cdd6f4", bg="#313244"
        ).pack()
        tk.Label(
            hdr, text="P&ID 마크업을 동일 Train 도면에 자동 복사",
            font=("Segoe UI", 9), fg="#a6adc8", bg="#313244"
        ).pack()

    def _build_body(self):
        body = tk.Frame(self, bg="#1e1e2e")
        body.pack(fill="both", expand=True, padx=14, pady=10)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=4)
        body.rowconfigure(0, weight=1)
        self._build_left(body)
        self._build_right(body)

    def _build_footer(self):
        footer = tk.Frame(self, bg="#181825", pady=8)
        footer.pack(fill="x", side="bottom")
        self.progress = ttk.Progressbar(footer, mode="determinate")
        self.progress.pack(fill="x", padx=14, pady=(0, 4))
        self.lbl_status = tk.Label(
            footer, text="대기 중",
            font=("Segoe UI", 9), fg="#a6adc8", bg="#181825"
        )
        self.lbl_status.pack()

    # ── 왼쪽 패널 ──────────────────────────────────────────

    def _build_left(self, parent):
        left = tk.Frame(parent, bg="#1e1e2e")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        left.rowconfigure(4, weight=1)   # 파일 목록 영역 확장
        left.columnconfigure(0, weight=1)

        # 폴더 설정 섹션
        self._sec(left, "📁 폴더 설정", row=0)

        folder_frm = tk.Frame(left, bg="#1e1e2e")
        folder_frm.grid(row=1, column=0, sticky="ew")
        folder_frm.columnconfigure(0, weight=1)

        self.lbl_src = self._folder_row(folder_frm, 0, "Source 폴더",
                                        self._select_source, "#a6e3a1",
                                        folder_cmd=self._select_source_folder)
        self.lbl_dst = self._folder_row(folder_frm, 1, "Target 폴더",
                                        self._select_target, "#89b4fa",
                                        folder_cmd=self._select_target_folder)
        self.lbl_out = self._folder_row(folder_frm, 2, "Output 폴더",
                                        self._select_output, "#f9e2af")

        # 파일 목록 섹션
        self._sec(left, "📄 파일 목록", row=2)

        # 탭으로 src / dst / out 파일 표시
        notebook = ttk.Notebook(left)
        notebook.grid(row=3, column=0, sticky="nsew", pady=(0, 6))
        left.rowconfigure(3, weight=1)

        nb_style = ttk.Style()
        nb_style.theme_use("clam")
        nb_style.configure("TNotebook",        background="#1e1e2e", borderwidth=0)
        nb_style.configure("TNotebook.Tab",    background="#313244", foreground="#a6adc8",
                           padding=[8, 4], font=("Segoe UI", 8))
        nb_style.map("TNotebook.Tab",
                     background=[("selected", "#45475a")],
                     foreground=[("selected", "#cdd6f4")])

        self.notebook = notebook
        self.list_src = self._file_listbox(notebook, "Source")
        self.list_dst = self._file_listbox(notebook, "Target")
        self.list_out = self._file_listbox(notebook, "Output")

        # 매핑 + 실행 섹션
        self._sec(left, "🔗 매핑 / 실행", row=4)
        left.rowconfigure(4, weight=0)

        btn_frm = tk.Frame(left, bg="#1e1e2e")
        btn_frm.grid(row=5, column=0, sticky="ew", pady=(0, 4))
        btn_frm.columnconfigure(0, weight=1)
        btn_frm.columnconfigure(0, weight=1)
        btn_frm.columnconfigure(1, weight=1)
        btn_frm.columnconfigure(2, weight=1)
        btn_frm.columnconfigure(3, weight=1)

        tk.Button(btn_frm, text="매핑 편집",
                  command=self._open_mapping,
                  **self._bkw("#7f849c")
                  ).grid(row=0, column=0, sticky="ew", padx=2)
        tk.Button(btn_frm, text="🗑 초기화",
                  command=self._reset_mapping,
                  **self._bkw("#585b70")
                  ).grid(row=1, column=0, sticky="ew", padx=2, pady=(2, 0))

        tk.Button(btn_frm, text="📂 Excel 매핑",
                  command=self._load_mapping_excel,
                  **self._bkw("#cba6f7", fg="#1e1e2e")
                  ).grid(row=0, column=1, sticky="ew", padx=2)

        tk.Button(btn_frm, text="🚀 실행",
                  command=self._start_run,
                  **self._bkw("#a6e3a1", fg="#1e1e2e", bold=True)
                  ).grid(row=0, column=2, sticky="ew", padx=2)

        tk.Button(btn_frm, text="⏹ 중지",
                  command=self._stop,
                  **self._bkw("#f38ba8", fg="#1e1e2e")
                  ).grid(row=0, column=3, sticky="ew", padx=2)

        self.lbl_mapping = tk.Label(
            left, text="매핑: 0 쌍",
            font=("Segoe UI", 9), fg="#a6adc8", bg="#1e1e2e"
        )
        self.lbl_mapping.grid(row=6, column=0, sticky="w")

        # 마크업 필터 섹션
        self._sec(left, "🎯 마크업 필터 (선택)", row=7)
        self._build_filter_section(left, row=8)

        # 위치 보정 섹션 (도면 레이아웃이 다른 경우)
        self._sec(left, "📐 도면 위치 보정 (선택)", row=9)
        self._build_position_section(left, row=10)

    def _build_filter_section(self, parent, row: int):
        frm = tk.Frame(parent, bg="#1e1e2e")
        frm.grid(row=row, column=0, sticky="ew", pady=(2, 0))
        frm.columnconfigure(1, weight=1)

        self.var_filter_enabled = tk.BooleanVar(value=False)
        tk.Checkbutton(
            frm, text="필터 사용 (특정 마크업만 붙여넣기)",
            variable=self.var_filter_enabled,
            font=("Segoe UI", 9), fg="#a6adc8", bg="#1e1e2e",
            selectcolor="#313244", activebackground="#1e1e2e",
            activeforeground="#cdd6f4"
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(2, 4))

        def _entry_row(r, label, width=24):
            tk.Label(frm, text=label, font=("Segoe UI", 8),
                     fg="#6c7086", bg="#1e1e2e", anchor="w", width=10
                     ).grid(row=r, column=0, sticky="w")
            e = tk.Entry(frm, width=width, bg="#181825", fg="#cdd6f4",
                          insertbackground="#cdd6f4", relief="flat")
            e.grid(row=r, column=1, sticky="ew", padx=4, pady=1)
            return e

        self.ent_date_limit = _entry_row(1, "기준 날짜")
        self.ent_author     = _entry_row(2, "작성자(사번)")

        tk.Label(
            frm,
            text="※ 기준 날짜: YYYY-MM-DD (예: 2026-05-08)\n"
                 "   이 날짜까지 작성된 마크업만 Target에 붙여넣습니다.\n"
                 "   작성자: 마크업 작성자(사번)와 정확히 일치해야 함\n"
                 "   (비워두면 해당 조건은 무시)",
            font=("Segoe UI", 8), fg="#6c7086", bg="#1e1e2e",
            justify="left", anchor="w"
        ).grid(row=3, column=0, columnspan=2, sticky="w", pady=(4, 2))

    def _build_position_section(self, parent, row: int):
        frm = tk.Frame(parent, bg="#1e1e2e")
        frm.grid(row=row, column=0, sticky="ew", pady=(2, 0))
        frm.columnconfigure(1, weight=1)

        self.var_pos_correction = tk.BooleanVar(value=False)
        tk.Checkbutton(
            frm, text="위치 보정 사용 (Source/Target 도면 레이아웃이 다른 경우)",
            variable=self.var_pos_correction,
            font=("Segoe UI", 9), fg="#a6adc8", bg="#1e1e2e",
            selectcolor="#313244", activebackground="#1e1e2e",
            activeforeground="#cdd6f4", wraplength=280, justify="left"
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(2, 4))

        tk.Label(
            frm,
            text="※ Source/Target에 마젠타색(255,0,255) 작은 마크업을 같은 순서로\n"
                 "   직접 찍어두면 그 위치를 기준점으로 우선 사용합니다(가장 정확).\n"
                 "   없으면 배관선번호/계측기/제어밸브/Symbol/도면 모서리를 자동으로\n"
                 "   찾아 기준점으로 사용 (수동 밸브 번호는 Train별로 달라져 제외).\n"
                 "   마크업마다 가까운 기준점 기준 국소 보정, Bluebeam 자동화 없이 직접 적용",
            font=("Segoe UI", 8), fg="#6c7086", bg="#1e1e2e",
            justify="left", anchor="w"
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(4, 2))

    def _file_listbox(self, notebook: ttk.Notebook, tab_name: str) -> tk.Listbox:
        """탭 내 스크롤 가능한 파일 목록 Listbox 생성"""
        frm = tk.Frame(notebook, bg="#181825")
        notebook.add(frm, text=tab_name)
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)

        lb = tk.Listbox(
            frm,
            bg="#181825", fg="#cdd6f4",
            selectbackground="#45475a",
            font=("Consolas", 8),
            bd=0, highlightthickness=0,
            activestyle="none",
            selectmode="extended"
        )
        lb.grid(row=0, column=0, sticky="nsew")

        sb = ttk.Scrollbar(frm, orient="vertical", command=lb.yview)
        sb.grid(row=0, column=1, sticky="ns")
        lb.configure(yscrollcommand=sb.set)
        return lb

    # ── 오른쪽 패널 (로그) — grid 전용 ────────────────────

    def _build_right(self, parent):
        right = tk.Frame(parent, bg="#1e1e2e")
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(2, weight=1)
        right.columnconfigure(0, weight=1)

        lbl = tk.Label(right, text="📋 작업 로그",
                       font=("Segoe UI", 10, "bold"),
                       fg="#cba6f7", bg="#1e1e2e", anchor="w")
        lbl.grid(row=0, column=0, sticky="ew", pady=(10, 0))

        sep = tk.Frame(right, bg="#45475a", height=1)
        sep.grid(row=1, column=0, sticky="ew", pady=(2, 6))

        log_frame = tk.Frame(right, bg="#181825", bd=1, relief="sunken")
        log_frame.grid(row=2, column=0, sticky="nsew")
        log_frame.rowconfigure(0, weight=1)
        log_frame.columnconfigure(0, weight=1)

        self.log_text = tk.Text(
            log_frame,
            bg="#181825", fg="#cdd6f4",
            font=("Consolas", 9),
            state="disabled", wrap="word",
            bd=0, padx=6, pady=6
        )
        self.log_text.grid(row=0, column=0, sticky="nsew")

        sb = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        sb.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=sb.set)

        btn_wrap = tk.Frame(right, bg="#1e1e2e")
        btn_wrap.grid(row=3, column=0, sticky="e", pady=(4, 0))
        tk.Button(btn_wrap, text="로그 지우기",
                  command=self._clear_log,
                  **self._bkw("#45475a", pady=3)).pack()

    # ── 공통 헬퍼 ──────────────────────────────────────────

    def _sec(self, parent, title: str, row: int):
        """grid 기반 섹션 제목 + 구분선"""
        tk.Label(parent, text=title,
                 font=("Segoe UI", 10, "bold"),
                 fg="#cba6f7", bg="#1e1e2e", anchor="w"
                 ).grid(row=row, column=0, sticky="ew", pady=(10, 0))

    def _folder_row(self, parent, row: int, label: str,
                    cmd, accent: str, folder_cmd=None) -> tk.Label:
        frm = tk.Frame(parent, bg="#1e1e2e")
        frm.grid(row=row, column=0, sticky="ew", pady=2)
        frm.columnconfigure(2, weight=1)

        tk.Button(frm, text=label, command=cmd, width=13,
                  **self._bkw(accent, fg="#1e1e2e")
                  ).grid(row=0, column=0, sticky="w")

        if folder_cmd is not None:
            tk.Button(frm, text="폴더 전체", command=folder_cmd, width=8,
                      **self._bkw("#6c7086", fg="#cdd6f4")
                      ).grid(row=0, column=1, sticky="w", padx=(4, 0))

        lbl = tk.Label(frm, text="(미설정)",
                       font=("Segoe UI", 8), fg="#6c7086",
                       bg="#1e1e2e", anchor="w", wraplength=220)
        lbl.grid(row=0, column=2, sticky="ew", padx=6)
        return lbl

    @staticmethod
    def _bkw(bg: str, fg: str = "#cdd6f4",
             bold: bool = False, pady: int = 5) -> dict:
        return dict(
            bg=bg, fg=fg, relief="flat",
            font=("Segoe UI", 9, "bold" if bold else "normal"),
            activebackground=bg, activeforeground=fg,
            cursor="hand2", padx=8, pady=pady, bd=0
        )

    # ── 폴더 선택 ──────────────────────────────────────────

    def _select_source(self):
        global src_folder, src_files
        files = filedialog.askopenfilenames(
            title="Source 폴더 - PDF 파일 선택 (Ctrl+A로 전체 선택)",
            filetypes=[("PDF 파일", "*.pdf")]
        )
        if not files:
            return
        src_folder = os.path.dirname(files[0])
        src_files = sorted(os.path.basename(f) for f in files)
        self.lbl_src.config(
            text=f"{os.path.basename(src_folder)}  ({len(src_files)}개)", fg="#a6adc8"
        )
        self._update_listbox(self.list_src, src_files)
        self.notebook.select(0)
        self._log(f"[Source] {src_folder}  ({len(src_files)}개)\n")
        for f in src_files:
            self._log(f"    - {f}\n")
        self._refresh_mapping_label()

    def _select_source_folder(self):
        global src_folder, src_files
        path = filedialog.askdirectory(title="Source 폴더 선택 (폴더 내 모든 PDF 사용)")
        if not path:
            return
        src_folder = path
        src_files = sorted(_list_pdfs(path))
        self.lbl_src.config(
            text=f"{os.path.basename(src_folder)}  ({len(src_files)}개)", fg="#a6adc8"
        )
        self._update_listbox(self.list_src, src_files)
        self.notebook.select(0)
        self._log(f"[Source] {src_folder}  ({len(src_files)}개)\n")
        for f in src_files:
            self._log(f"    - {f}\n")
        self._refresh_mapping_label()

    def _select_target_folder(self):
        global dst_folder, dst_files
        path = filedialog.askdirectory(title="Target 폴더 선택 (폴더 내 모든 PDF 사용)")
        if not path:
            return
        dst_folder = path
        dst_files = sorted(_list_pdfs(path))
        self.lbl_dst.config(
            text=f"{os.path.basename(dst_folder)}  ({len(dst_files)}개)", fg="#a6adc8"
        )
        self._update_listbox(self.list_dst, dst_files)
        self.notebook.select(1)
        self._log(f"[Target] {dst_folder}  ({len(dst_files)}개)\n")
        for f in dst_files:
            self._log(f"    - {f}\n")
        self._refresh_mapping_label()

    def _select_target(self):
        global dst_folder, dst_files
        files = filedialog.askopenfilenames(
            title="Target 폴더 - PDF 파일 선택 (Ctrl+A로 전체 선택)",
            filetypes=[("PDF 파일", "*.pdf")]
        )
        if not files:
            return
        dst_folder = os.path.dirname(files[0])
        dst_files = sorted(os.path.basename(f) for f in files)
        self.lbl_dst.config(
            text=f"{os.path.basename(dst_folder)}  ({len(dst_files)}개)", fg="#a6adc8"
        )
        self._update_listbox(self.list_dst, dst_files)
        self.notebook.select(1)
        self._log(f"[Target] {dst_folder}  ({len(dst_files)}개)\n")
        for f in dst_files:
            self._log(f"    - {f}\n")
        self._refresh_mapping_label()

    def _select_output(self):
        global output_folder
        path = filedialog.askdirectory(title="Output 폴더 선택")
        if not path:
            return
        output_folder = path
        # Output 폴더의 기존 PDF 표시
        out_files = sorted(_list_pdfs(path))
        self.lbl_out.config(text=os.path.basename(path), fg="#a6adc8")
        self._update_listbox(self.list_out, out_files)
        self.notebook.select(2)
        self._log(f"[Output] {path}  ({len(out_files)}개)\n")
        for f in out_files:
            self._log(f"    - {f}\n")

    def _update_listbox(self, lb: tk.Listbox, files: list):
        lb.delete(0, "end")
        for i, f in enumerate(files, 1):
            lb.insert("end", f"  {i:>3}.  {f}")
        # 짝수 행 색상 구분
        for i in range(0, len(files), 2):
            lb.itemconfig(i, bg="#1e1e2e")
        for i in range(1, len(files), 2):
            lb.itemconfig(i, bg="#181825")

    # ── Excel 매핑 가져오기 ─────────────────────────────────

    def _load_mapping_excel(self):
        """Excel 파일로 매핑 가져오기.

        Excel 형식:
          1행: 헤더 (무시)
          2행: 폴더 경로행 — A2=Source 폴더, B2~=각 Train의 Target 폴더
               (셀 값에 경로 구분자 \\ 또는 / 포함 시 경로행으로 자동 인식)
          3행~: 파일명행 — A열=Source 파일명, B~N열=Train별 Target 파일명
               (Source 하나에 Target 여러 개 → 각각 별도 매핑 쌍으로 확장)
        """
        global mapping, src_files, dst_files, src_folder, dst_folder
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror(
                "openpyxl 필요",
                "Excel 매핑 기능에는 openpyxl이 필요합니다.\n"
                "터미널에서 'pip install openpyxl' 실행 후 다시 시도하세요."
            )
            return

        path = filedialog.askopenfilename(
            title="매핑 Excel 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")]
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("파일 오류", f"Excel 파일을 열 수 없습니다:\n{e}")
            return

        def _is_path(val):
            s = str(val or "").strip()
            return bool(s) and ("\\" in s or "/" in s or (len(s) > 2 and s[1] == ":"))

        def _as_str(val):
            s = str(val or "").strip()
            if not s or s.lower() == "none":
                return ""
            return s

        def _ensure_pdf(name):
            return name if name.lower().endswith(".pdf") else name + ".pdf"

        # 모든 행 읽기 (헤더 1행 스킵)
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not all_rows:
            messagebox.showwarning("내용 없음", "Excel 파일에 데이터가 없습니다.")
            return

        # 폴더 경로 행 감지: 첫 데이터행(2행)에 경로 구분자가 있으면 폴더 행으로 간주
        col_dirs = []   # 열별 Target 폴더 경로 (col 0 = Source 폴더)
        data_start = 0
        first = all_rows[0]
        if any(_is_path(c) for c in first if c):
            col_dirs = [_as_str(c) for c in first]
            data_start = 1
            # Source 폴더(A열 경로)가 있으면 GUI lbl_src 갱신
            if col_dirs[0]:
                self._log(f"[Excel 매핑] Source 폴더: {col_dirs[0]}\n")

        n_targets = (ws.max_column - 1)  # A열 제외 나머지 = Target 열 수

        new_mapping = []
        for row in all_rows[data_start:]:
            src_name = _as_str(row[0])
            if not src_name:
                continue
            src_name = _ensure_pdf(src_name)
            s_dir = col_dirs[0] if col_dirs else src_folder

            for col_i in range(1, len(row)):
                dst_name = _as_str(row[col_i])
                if not dst_name:
                    continue
                dst_name = _ensure_pdf(dst_name)
                d_dir = col_dirs[col_i] if col_dirs and col_i < len(col_dirs) else dst_folder
                new_mapping.append((src_name, dst_name, s_dir, d_dir))

        if not new_mapping:
            messagebox.showwarning("매핑 없음",
                "Excel에서 유효한 매핑을 찾지 못했습니다.\n\n"
                "형식:\n"
                "  1행: 헤더 (무시)\n"
                "  2행: 폴더 경로 (선택, 경로 구분자 포함 시 자동 인식)\n"
                "  3행~: A열=Source 파일명, B~열=Train별 Target 파일명")
            return

        mapping   = new_mapping
        src_files = list(dict.fromkeys(r[0] for r in mapping))
        dst_files = list(dict.fromkeys(r[1] for r in mapping))

        self._update_listbox(self.list_src, src_files)
        self._update_listbox(self.list_dst, dst_files)
        self._refresh_mapping_label()

        # 폴더 라벨 업데이트
        if col_dirs and col_dirs[0]:
            src_folder = col_dirs[0]
            self.lbl_src.config(text=os.path.basename(src_folder.rstrip("/\\")) or src_folder,
                                fg="#a6e3a1")
        dst_dirs = [d for _, _, _, d in mapping if d]
        if dst_dirs:
            dst_folder = dst_dirs[0]
            unique_dst_dirs = list(dict.fromkeys(d for _, _, _, d in mapping if d))
            label = unique_dst_dirs[0] if len(unique_dst_dirs) == 1 else f"{len(unique_dst_dirs)}개 폴더"
            self.lbl_dst.config(text=os.path.basename(label.rstrip("/\\")) or label,
                                fg="#a6e3a1")

        self._log(f"[Excel 매핑] {len(mapping)}개 매핑 로드 ← {os.path.basename(path)}\n")
        for s, d, sd, dd in mapping:
            self._log(f"    [{os.path.basename(sd or src_folder)}] {s}  →  [{os.path.basename(dd or dst_folder)}] {d}\n")

        messagebox.showinfo("Excel 매핑 완료",
                            f"{len(mapping)}개 매핑을 가져왔습니다.\n"
                            f"'매핑 편집'에서 확인/수정 가능합니다.")

    # ── 매핑 편집 창 ───────────────────────────────────────

    def _reset_mapping(self):
        global mapping, src_files, dst_files
        if not mapping and not src_files and not dst_files:
            return
        if messagebox.askyesno("매핑 초기화", "현재 매핑을 모두 지울까요?"):
            mapping = []
            src_files.clear()
            dst_files.clear()
            self.list_src.delete(0, tk.END)
            self.list_dst.delete(0, tk.END)
            self.lbl_mapping.config(text="매핑: 0 쌍")
            self._log("[매핑 초기화]\n")

    def _open_mapping(self):
        global mapping
        if not src_files or not dst_files:
            if not mapping:
                messagebox.showwarning("폴더 미설정", "Source / Target 폴더를 먼저 선택하거나 Excel 매핑을 불러오세요.")
                return

        if not mapping:
            sel_src = [src_files[i] for i in self.list_src.curselection()]
            sel_dst = [dst_files[i] for i in self.list_dst.curselection()]
            use_src = sel_src or src_files
            use_dst = sel_dst or dst_files
            if len(use_src) != len(use_dst):
                messagebox.showwarning(
                    "파일 수 불일치",
                    f"Source {len(use_src)}개 ≠ Target {len(use_dst)}개\n"
                    "매핑 창에서 직접 삭제/조정하세요."
                )
            mapping = [(s, d, src_folder, dst_folder) for s, d in zip(use_src, use_dst)]

        win = tk.Toplevel(self)
        win.title(f"파일 매핑 편집  ({len(mapping)}쌍)")
        win.geometry("1050x520")
        win.configure(bg="#1e1e2e")
        win.grab_set()

        tk.Label(win, text="우클릭 → 선택 행 삭제  |  확정 버튼을 눌러야 반영됩니다",
                 font=("Segoe UI", 9), fg="#a6adc8", bg="#1e1e2e"
                 ).pack(pady=(8, 2))

        style = ttk.Style(win)
        style.theme_use("clam")
        style.configure("Map.Treeview",
                        background="#181825", fieldbackground="#181825",
                        foreground="#cdd6f4", rowheight=24,
                        font=("Consolas", 9))
        style.configure("Map.Treeview.Heading",
                        background="#313244", foreground="#cba6f7",
                        font=("Segoe UI", 9, "bold"))

        frm = tk.Frame(win, bg="#1e1e2e")
        frm.pack(fill="both", expand=True, padx=12, pady=4)

        tree = ttk.Treeview(frm, style="Map.Treeview",
                            columns=("#", "src", "src_dir", "dst", "dst_dir"),
                            show="headings")
        tree.heading("#",       text="#",           anchor="center")
        tree.heading("src",     text="Source 파일명")
        tree.heading("src_dir", text="Source 폴더")
        tree.heading("dst",     text="Target 파일명")
        tree.heading("dst_dir", text="Target 폴더")
        tree.column("#",       width=36,  anchor="center", stretch=False)
        tree.column("src",     width=260)
        tree.column("src_dir", width=180)
        tree.column("dst",     width=260)
        tree.column("dst_dir", width=180)

        vsb = ttk.Scrollbar(frm, orient="vertical",   command=tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)

        for i, row in enumerate(mapping, 1):
            s  = row[0]; d  = row[1]
            sd = row[2] if len(row) > 2 else ""
            dd = row[3] if len(row) > 3 else ""
            tree.insert("", "end", values=(i, s, sd, d, dd))

        ctx = tk.Menu(win, tearoff=0, bg="#313244", fg="#cdd6f4",
                      activebackground="#45475a")

        def _delete():
            for item in tree.selection():
                tree.delete(item)
            for i, row in enumerate(tree.get_children(), 1):
                vals = list(tree.item(row)["values"])
                vals[0] = i
                tree.item(row, values=vals)

        ctx.add_command(label="선택 행 삭제", command=_delete)
        tree.bind("<Button-3>", lambda e: ctx.post(e.x_root, e.y_root))

        btn_frm = tk.Frame(win, bg="#1e1e2e")
        btn_frm.pack(fill="x", padx=12, pady=8)

        def _confirm():
            global mapping
            mapping = [
                (tree.item(r)["values"][1],
                 tree.item(r)["values"][3],
                 tree.item(r)["values"][2],
                 tree.item(r)["values"][4])
                for r in tree.get_children()
            ]
            self._refresh_mapping_label()
            self._log(f"[매핑 확정] {len(mapping)}쌍\n")
            win.destroy()

        tk.Button(btn_frm, text="✔  매핑 확정",
                  command=_confirm,
                  **self._bkw("#a6e3a1", fg="#1e1e2e", bold=True)
                  ).pack(side="right", padx=4)
        tk.Button(btn_frm, text="✖  닫기",
                  command=win.destroy,
                  **self._bkw("#f38ba8", fg="#1e1e2e")
                  ).pack(side="right")

    def _refresh_mapping_label(self):
        n = len(mapping)
        self.lbl_mapping.config(text=f"매핑: {n} 쌍")

    # ── 실행 / 중지 ────────────────────────────────────────

    def _start_run(self):
        global stop_flag, filter_enabled, filter_color
        global filter_date_limit, filter_author
        global pos_correction_enabled
        if not mapping:
            messagebox.showwarning("매핑 없음", "먼저 '매핑 편집'에서 확정하세요.")
            return
        if not output_folder:
            messagebox.showwarning("Output 미설정", "Output 폴더를 설정하세요.")
            return

        filter_enabled   = self.var_filter_enabled.get()
        filter_color     = ""
        filter_date_limit = _normalize_date_input(self.ent_date_limit.get())
        filter_author    = self.ent_author.get().strip()

        pos_correction_enabled = self.var_pos_correction.get()

        if filter_enabled or pos_correction_enabled:
            if fitz is None:
                messagebox.showerror(
                    "PyMuPDF 필요",
                    "마크업 필터 / 위치 보정 기능에는 PyMuPDF가 필요합니다.\n"
                    "터미널에서 'pip install PyMuPDF' 실행 후 다시 시도하세요."
                )
                return

        stop_flag = False
        self._set_status("실행 중…", "#a6e3a1")
        self.progress["maximum"] = len(mapping)
        self.progress["value"]   = 0
        threading.Thread(target=self._run_worker, daemon=True).start()

    def _stop(self):
        global stop_flag
        stop_flag = True
        self._set_status("중지 요청됨…", "#f9e2af")

    def _run_worker(self):
        report_rows = []
        try:
            self._run_worker_inner(report_rows)
        except Exception:
            err = traceback.format_exc()
            self._log(f"\n💥 예기치 않은 오류로 작업 중단:\n{err}\n")
            self._set_status("오류로 중단 ❌", "#f38ba8")
            self._write_report(report_rows)
            try:
                close_bluebeam_app()
            except Exception:
                pass

    def _run_worker_inner(self, report_rows):
        filter_settings = None
        if filter_enabled:
            filter_settings = dict(
                color=filter_color,
                date_limit=filter_date_limit,
                author=filter_author,
            )

        # Source별로 그룹핑: {(src_path): [(dst_path, out_path, src_name, dst_name), ...]}
        from collections import OrderedDict
        groups = OrderedDict()
        for row in mapping:
            src, dst = row[0], row[1]
            s_dir = row[2] if len(row) > 2 and row[2] else src_folder
            d_dir = row[3] if len(row) > 3 and row[3] else dst_folder
            src_path = os.path.join(s_dir, src)
            dst_path = os.path.join(d_dir, dst)
            out_path = os.path.join(output_folder, dst)
            groups.setdefault(src_path, []).append((dst_path, out_path, src, dst))

        total = len(mapping)
        done = 0

        for src_path, targets in groups.items():
            src_name = os.path.basename(src_path)
            n = len(targets)
            self._log(f"\n▶ Source: {src_name}  ({n}개 Target)\n")
            self._set_status(f"처리 중: {done + 1}~{done + n} / {total}")

            if stop_flag:
                self._log("\n⏹ 사용자에 의해 중지됨\n")
                self._set_status("중지됨", "#f38ba8")
                self._write_report(report_rows)
                return

            if pos_correction_enabled:
                done = self._process_group_with_position_correction(
                    src_path, targets, filter_settings, report_rows, done, total
                )
                continue

            # Source 열기 → 복사 (한 번만)
            start_src = time.time()
            try:
                filter_kept, filter_removed = _open_and_copy_source(
                    src_path, filter_settings, self._log
                )
            except StoppedError:
                self._log("\n⏹ 사용자에 의해 중지됨\n")
                self._set_status("중지됨", "#f38ba8")
                self._write_report(report_rows)
                return
            except Exception:
                err = traceback.format_exc()
                self._log(f"  ⚠ Source 열기 오류:\n{err}\n")
                for _, _, src_name2, dst_name in targets:
                    report_rows.append({
                        "source": src_name2, "target": dst_name, "output": dst_name,
                        "status": "오류(Source)", "filter_kept": "", "filter_removed": "",
                        "elapsed_sec": round(time.time() - start_src, 1),
                        "error": str(err).splitlines()[-1] if err else "",
                    })
                done += n
                self.after(0, lambda v=done: self.progress.configure(value=v))
                continue

            # 각 Target에 붙여넣기
            for dst_path, out_path, src_name2, dst_name in targets:
                done += 1
                self._log(f"  [{done}/{total}] → {dst_name}\n")
                start_time = time.time()
                try:
                    _paste_to_target(dst_path, out_path, self._log, lambda: stop_flag)
                    elapsed = time.time() - start_time
                    report_rows.append({
                        "source": src_name2, "target": dst_name, "output": dst_name,
                        "status": "완료",
                        "filter_kept": filter_kept,
                        "filter_removed": filter_removed,
                        "elapsed_sec": round(elapsed, 1),
                        "error": "",
                    })
                except StoppedError:
                    self._log("\n⏹ 사용자에 의해 중지됨\n")
                    self._set_status("중지됨", "#f38ba8")
                    self._write_report(report_rows)
                    return
                except Exception:
                    err = traceback.format_exc()
                    self._log(f"  ⚠ 오류:\n{err}\n")
                    report_rows.append({
                        "source": src_name2, "target": dst_name, "output": dst_name,
                        "status": "오류", "filter_kept": "", "filter_removed": "",
                        "elapsed_sec": round(time.time() - start_time, 1),
                        "error": str(err).splitlines()[-1] if err else "",
                    })

                self.after(0, lambda v=done: self.progress.configure(value=v))
                out_files = sorted(_list_pdfs(output_folder))
                self.after(0, lambda fl=out_files: self._update_listbox(self.list_out, fl))

        self._write_report(report_rows)

        # 모든 작업 완료 후 Bluebeam 완전 종료 (위치 보정 모드는 Bluebeam을
        # 전혀 쓰지 않으므로 건너뜀)
        if not pos_correction_enabled:
            self._log("\n[마무리] Bluebeam 종료 중…\n")
            close_bluebeam_app()

        self._log("✅ 모든 작업 완료!\n")
        self._set_status("완료 ✅", "#a6e3a1")

        if open_folder(output_folder):
            self._log(f"📂 Output 폴더 열기: {output_folder}\n")
        else:
            self._log(f"⚠ Output 폴더를 열지 못했습니다: {output_folder}\n")

    def _process_group_with_position_correction(self, src_path, targets,
                                                 filter_settings, report_rows,
                                                 done, total):
        """위치 보정 모드: Bluebeam 자동화 없이 PyMuPDF로 직접 좌표 보정하여 복사.
        done(처리 완료 개수)을 갱신해서 반환한다."""
        src_name = os.path.basename(src_path)
        self._log(f"\n▶ Source: {src_name}  ({len(targets)}개 Target) [위치 보정 모드]\n")

        open_src = src_path
        tmp_to_delete = None
        if filter_settings:
            try:
                open_src, kept, removed = build_filtered_copy(
                    src_path, "",
                    filter_settings.get("date_limit", ""),
                    filter_settings.get("author", ""),
                    log_fn=self._log,
                )
                if open_src != src_path:
                    tmp_to_delete = open_src
            except Exception:
                err = traceback.format_exc()
                self._log(f"  ⚠ 필터 적용 오류:\n{err}\n")
                for _, _, src_name2, dst_name in targets:
                    done += 1
                    report_rows.append({
                        "source": src_name2, "target": dst_name, "output": dst_name,
                        "status": "오류(필터)", "filter_kept": "", "filter_removed": "",
                        "elapsed_sec": 0, "error": str(err).splitlines()[-1] if err else "",
                    })
                self.after(0, lambda v=done: self.progress.configure(value=v))
                return done

        for dst_path, out_path, src_name2, dst_name in targets:
            done += 1
            self._log(f"  [{done}/{total}] → {dst_name}\n")
            start_time = time.time()
            try:
                copied, skipped = run_position_correction_isolated(
                    open_src, dst_path, out_path, self._log
                )
                self._log(f"  → 복사 {copied}개 / 스킵 {skipped}개\n")
                report_rows.append({
                    "source": src_name2, "target": dst_name, "output": dst_name,
                    "status": "완료(위치보정)",
                    "filter_kept": copied, "filter_removed": skipped,
                    "elapsed_sec": round(time.time() - start_time, 1),
                    "error": "",
                })
            except Exception:
                err = traceback.format_exc()
                self._log(f"  ⚠ 오류:\n{err}\n")
                report_rows.append({
                    "source": src_name2, "target": dst_name, "output": dst_name,
                    "status": "오류", "filter_kept": "", "filter_removed": "",
                    "elapsed_sec": round(time.time() - start_time, 1),
                    "error": str(err).splitlines()[-1] if err else "",
                })

            self.after(0, lambda v=done: self.progress.configure(value=v))
            out_files = sorted(_list_pdfs(output_folder))
            self.after(0, lambda fl=out_files: self._update_listbox(self.list_out, fl))

        if tmp_to_delete:
            try:
                os.remove(tmp_to_delete)
            except OSError:
                pass

        return done

    def _write_report(self, report_rows):
        """처리 결과를 Output 폴더에 CSV로 저장"""
        if not report_rows or not output_folder:
            return
        report_path = os.path.join(output_folder, "처리결과.csv")
        try:
            with open(report_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=[
                    "source", "target", "output", "status",
                    "filter_kept", "filter_removed", "elapsed_sec", "error",
                ])
                writer.writeheader()
                writer.writerows(report_rows)
            self._log(f"\n📄 처리 결과 리포트 저장: {report_path}\n")
        except OSError:
            self._log(f"\n⚠ 처리 결과 리포트 저장 실패: {report_path}\n")

    # ── 로그 헬퍼 ──────────────────────────────────────────

    def _log(self, msg: str):
        def _write():
            self.log_text.configure(state="normal")
            self.log_text.insert("end", msg)
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
        self.after(0, _write)

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _set_status(self, msg: str, color: str = "#a6adc8"):
        self.after(0, lambda: self.lbl_status.config(text=msg, fg=color))


# ══════════════════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════════════════

if __name__ == "__main__":
    mp.freeze_support()
    app = App()
    app.mainloop()
